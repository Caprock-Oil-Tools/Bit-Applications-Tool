"""
Cutlet Geometry Engine — Computational replication of the AutoCAD MASSPROP method.

Replicates the VBA macro logic from '2176r06 Min Engagement v8.01.xlsm':
  1. Reads cutter geometry from an ME file (Assy.Model sheet)
  2. Constructs ellipses on the Revolved Z Projection plane
  3. Creates 3 revolutions (previous, current, next) offset by IPR
  4. For each Rev 2 cutter, subtracts all lower-indexed ellipses
  5. Computes cutlet area and centroid

Sequencing matches the VBA: blade order from spreadsheet, lower indices mask higher.
"""

import math
import numpy as np
from shapely.geometry import Polygon, box
from shapely.affinity import rotate, translate
from shapely import ops
import openpyxl


# ---------------------------------------------------------------------------
# Ellipse polygon construction (matches Ellipse.cls parameterization)
# ---------------------------------------------------------------------------

def make_ellipse_polygon(xc, yc, major, minor, tilt_deg, n_points=360):
    """
    Create a Shapely polygon approximating a rotated ellipse.

    Matches the VBA Ellipse.cls exactly:
      - major = semi-major axis (PDC radius from column H)
      - minor = major * cos(rake)
      - tilt_deg = negated tilt from column AN
      - Center at (xc, yc) where xc = -radial, yc = Z_drilling

    The ellipse is constructed in local coords then rotated by tilt and translated.
    """
    angles = np.linspace(0, 2 * math.pi, n_points, endpoint=False)
    # Local ellipse points (before rotation)
    xs = major * np.cos(angles)
    ys = minor * np.sin(angles)

    # Rotate by tilt angle
    tilt_rad = math.radians(tilt_deg)
    cos_t = math.cos(tilt_rad)
    sin_t = math.sin(tilt_rad)
    x_rot = xs * cos_t - ys * sin_t + xc
    y_rot = xs * sin_t + ys * cos_t + yc

    coords = list(zip(x_rot, y_rot))
    coords.append(coords[0])  # close the ring
    return Polygon(coords)


# ---------------------------------------------------------------------------
# Read cutter data from ME file
# ---------------------------------------------------------------------------

def read_cutter_data_from_me(filepath):
    """
    Read cutter geometry from an ME XLSM file's Assy.Model sheet.

    Returns list of dicts with keys:
        name, radial, z_stat, z_drill, major, tilt, rake, zpolar

    Uses the same Settings-sheet column mappings as the VBA StartUp_Module.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)

    # Read settings
    ws_settings = wb['Settings']
    start_row = int(ws_settings['C9'].value)
    name_col = ws_settings['B13'].value       # J
    x_col = ws_settings['C13'].value          # X (radial position)
    z_stat_col = ws_settings['D13'].value     # Y (stationary Z)
    z_drill_col = ws_settings['D14'].value    # AT (drilling Z with helical offset)
    rcutter_col = ws_settings['E13'].value    # H (PDC radius = semi-major)
    tilt_col = ws_settings['F13'].value       # AN
    rake_col = ws_settings['G13'].value       # AO
    zpolar_col = ws_settings['I13'].value     # Q

    ipr_row = int(ws_settings['D18'].value)
    ipr_col = ws_settings['C18'].value
    gage_radius = float(ws_settings['C25'].value)

    ws = wb['Assy.Model']

    # Read IPR
    ipr_cell = ws.cell(row=ipr_row, column=openpyxl.utils.column_index_from_string(ipr_col))
    ipr = float(ipr_cell.value)

    # Helper to get column index
    def col_idx(letter):
        return openpyxl.utils.column_index_from_string(letter)

    cutters = []
    row = start_row
    while True:
        name_val = ws.cell(row=row, column=col_idx(name_col)).value
        if name_val is None:
            break
        try:
            name_float = float(name_val)
        except (ValueError, TypeError):
            break

        radial = ws.cell(row=row, column=col_idx(x_col)).value
        z_stat = ws.cell(row=row, column=col_idx(z_stat_col)).value
        z_drill = ws.cell(row=row, column=col_idx(z_drill_col)).value
        major = ws.cell(row=row, column=col_idx(rcutter_col)).value
        tilt = ws.cell(row=row, column=col_idx(tilt_col)).value
        rake = ws.cell(row=row, column=col_idx(rake_col)).value
        zpolar = ws.cell(row=row, column=col_idx(zpolar_col)).value

        # Read element type from column K to identify knuckles (CPS elements)
        element = ws.cell(row=row, column=col_idx('K')).value
        element_str = str(element).upper() if element else ''
        is_knuckle = 'CPS' in element_str

        if radial is None or major is None:
            break

        cutters.append({
            'name': name_float,
            'radial': float(radial),
            'z_stat': float(z_stat) if z_stat is not None else 0.0,
            'z_drill': float(z_drill) if z_drill is not None else float(z_stat) if z_stat else 0.0,
            'major': float(major),
            'tilt': float(tilt) if tilt is not None else 0.0,
            'rake': float(rake) if rake is not None else 0.0,
            'zpolar': float(zpolar) if zpolar is not None else 0.0,
            'is_knuckle': is_knuckle,
        })
        row += 1

    wb.close()
    return cutters, ipr, gage_radius


def read_massprop_ground_truth(filepath):
    """Read MassProp sheet ground truth for validation."""
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb['MassProp']

    truth = []
    for row in range(9, 50):
        name = ws.cell(row=row, column=2).value
        cx = ws.cell(row=row, column=3).value
        cy = ws.cell(row=row, column=4).value
        area = ws.cell(row=row, column=5).value
        if name is None:
            break
        truth.append({
            'name': float(name),
            'centroid_x': float(cx),
            'centroid_y': float(cy),
            'area': float(area),
        })

    wb.close()
    return truth


# ---------------------------------------------------------------------------
# Core cutlet computation
# ---------------------------------------------------------------------------

def _filter_non_drilling_cutters(cutters):
    """
    Remove back-reamers and wear pad elements from the cutter list.

    Back-reamers are at the opposite end of the bit, past wear pads.
    Detection: largest Z gap > 1 inch separates drilling from non-drilling cutters.
    """
    if len(cutters) < 2:
        return cutters

    sorted_by_z = sorted(cutters, key=lambda c: c['z_drill'])
    max_gap = 0
    max_gap_idx = -1
    for i in range(1, len(sorted_by_z)):
        gap = sorted_by_z[i]['z_drill'] - sorted_by_z[i - 1]['z_drill']
        if gap > max_gap:
            max_gap = gap
            max_gap_idx = i

    if max_gap > 1.0:
        drilling_cutters = sorted_by_z[max_gap_idx:]
        removed = sorted_by_z[:max_gap_idx]
        print(f"  Excluded {len(removed)} non-drilling cutters (back-reamers/wear pads)")
        return drilling_cutters

    return cutters


def compute_cutlets(cutters, ipr, gage_radius, n_ellipse_points=720):
    """
    Compute cutlet area and centroid for each cutter.

    Back-reamers and wear pad elements are excluded before computation.

    Based on the VBA macro flow, optimized to 2 revolutions:
      1. Build ellipses for all cutters (Rev 1 — original positions)
      2. Create Rev 2 copies shifted by IPR (these get measured)
      3. For each Rev 2 ellipse, subtract all lower-indexed ellipses
      4. Clip to gauge line (x in [-gage_radius, 0])
      5. Compute centroid, area, and perimeter

    Rev 3 (shifted by 2*IPR) is NOT needed: Rev 3 indices are higher than
    Rev 2, so they never mask Rev 2 cutlets. Rev 3 was only a visual bottom
    boundary in AutoCAD.

    Returns list of dicts with keys: name, centroid_x, centroid_y, area, perimeter
    """
    cutters = _filter_non_drilling_cutters(cutters)
    N = len(cutters)

    # Build all ellipse data: (name, xc, yc, major, minor, tilt, revolution, massprop_flag)
    all_ellipses = []

    # --- Revolution 1: original positions ---
    for c in cutters:
        xc = -c['radial']  # NEGATE radial (VBA: X * -1)
        yc = c['z_drill']
        major = c['major']
        tilt = -c['tilt']  # NEGATE tilt (VBA: tilt * -1)
        rake = c['rake']
        minor = major * math.cos(math.radians(rake))
        all_ellipses.append({
            'name': c['name'],
            'xc': xc, 'yc': yc,
            'major': major, 'minor': minor,
            'tilt': tilt,
            'rev': 1, 'massprop': False,
        })

    # --- Revolution 2: shifted by IPR, MassProp=True ---
    # VBA copies indices 0..N-2 (skips last cutter)
    for i in range(N - 1):
        e = all_ellipses[i]
        all_ellipses.append({
            'name': e['name'],
            'xc': e['xc'],
            'yc': e['yc'] + ipr,  # shiftDown(-IPR) => Yc = Yc - (-IPR) = Yc + IPR
            'major': e['major'], 'minor': e['minor'],
            'tilt': e['tilt'],
            'rev': 2, 'massprop': True,
        })

    # Rev 3 is NOT needed for cutlet computation.
    # Rev 3 ellipses have HIGHER indices than Rev 2, so they never mask Rev 2.
    # They were only used as a visual bottom boundary in the AutoCAD drawing.
    # Removing Rev 3 cuts total ellipses by ~33% and speeds computation by ~55%.

    total = len(all_ellipses)
    print(f"  Total ellipses: {total} ({N} cutters × 2 revs, minus 1 for last cutter)")

    # --- Build Shapely polygons ---
    polygons = []
    for e in all_ellipses:
        poly = make_ellipse_polygon(
            e['xc'], e['yc'], e['major'], e['minor'], e['tilt'],
            n_points=n_ellipse_points
        )
        polygons.append(poly)

    # --- Gauge clip box: x in [-gage_radius, 0], y generous ---
    # Find Y range for clip box
    all_y = [e['yc'] for e in all_ellipses]
    y_min = min(all_y) - 2.0
    y_max = max(all_y) + 2.0
    clip_box = box(-gage_radius, y_min, 0, y_max)

    # --- Compute cutlets for Rev 2 ellipses ---
    results = []

    for idx in range(total):
        if not all_ellipses[idx]['massprop']:
            continue

        # Start with this ellipse's polygon
        cutlet = polygons[idx]

        # Subtract all lower-indexed ellipses (the masking rule)
        for mask_idx in range(idx):
            if not cutlet.is_empty:
                cutlet = cutlet.difference(polygons[mask_idx])

        # Clip to gauge line
        if not cutlet.is_empty:
            cutlet = cutlet.intersection(clip_box)

        # Compute area and centroid
        if cutlet.is_empty or cutlet.area < 1e-10:
            continue

        area = cutlet.area
        centroid = cutlet.centroid

        # Centroid X is negated back to positive radial (VBA: CX = -1 * CX)
        centroid_x = -centroid.x
        centroid_y = centroid.y

        perimeter = cutlet.length

        results.append({
            'name': all_ellipses[idx]['name'],
            'centroid_x': round(centroid_x, 4),
            'centroid_y': round(centroid_y, 4),
            'area': round(area, 4),
            'perimeter': round(perimeter, 4),
        })

    # Sort by centroid_x ascending (inner → outer), matching VBA ProcessMassResults
    results.sort(key=lambda r: r['centroid_x'])

    return results


def compute_cutlets_from_raw(cutter_list, ipr=0.25, gage_radius=3.9375,
                             min_area=0.001, n_ellipse_points=720):
    """
    Compute cutlets from a list of raw cutter dicts (no ME file needed).

    Each dict must have: name, radial, z_drill, major, tilt, rake
    Optional: z_stat, zpolar

    This is the entry point for ME files that lack MassProp data.
    Returns list of dicts: name, centroid_x, centroid_y, area
    """
    results = compute_cutlets(cutter_list, ipr, gage_radius, n_ellipse_points)
    # Filter out tiny cutlets (gauge cutters with near-zero engagement)
    results = [r for r in results if r['area'] >= min_area]
    return results


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------

def validate_against_ground_truth(computed, truth):
    """Compare computed cutlets against MassProp ground truth."""
    # Build lookup by name
    truth_by_name = {t['name']: t for t in truth}
    computed_by_name = {c['name']: c for c in computed}

    print(f"\n{'Name':<8} {'GT_CX':>8} {'Comp_CX':>8} {'dCX':>7} "
          f"{'GT_CY':>8} {'Comp_CY':>8} {'dCY':>7} "
          f"{'GT_Area':>8} {'Comp_A':>8} {'%Err':>7}")
    print("-" * 90)

    total_area_err = 0
    total_cx_err = 0
    total_cy_err = 0
    matched = 0

    for name in sorted(truth_by_name.keys()):
        gt = truth_by_name[name]
        if name not in computed_by_name:
            print(f"{name:<8} {'MISSING':>60}")
            continue

        comp = computed_by_name[name]
        dcx = comp['centroid_x'] - gt['centroid_x']
        dcy = comp['centroid_y'] - gt['centroid_y']

        if gt['area'] > 0:
            area_pct = ((comp['area'] - gt['area']) / gt['area']) * 100
        else:
            area_pct = 0

        total_area_err += abs(area_pct)
        total_cx_err += abs(dcx)
        total_cy_err += abs(dcy)
        matched += 1

        print(f"{name:<8} {gt['centroid_x']:>8.4f} {comp['centroid_x']:>8.4f} {dcx:>+7.4f} "
              f"{gt['centroid_y']:>8.4f} {comp['centroid_y']:>8.4f} {dcy:>+7.4f} "
              f"{gt['area']:>8.4f} {comp['area']:>8.4f} {area_pct:>+6.1f}%")

    if matched > 0:
        print("-" * 90)
        print(f"{'AVG':.<8} {'':>8} {'':>8} {total_cx_err/matched:>7.4f} "
              f"{'':>8} {'':>8} {total_cy_err/matched:>7.4f} "
              f"{'':>8} {'':>8} {total_area_err/matched:>6.1f}%")

    # Count how many computed cutters are missing from ground truth
    extra = set(computed_by_name.keys()) - set(truth_by_name.keys())
    missing = set(truth_by_name.keys()) - set(computed_by_name.keys())
    if extra:
        print(f"\nExtra computed (not in GT): {sorted(extra)}")
    if missing:
        print(f"\nMissing computed (in GT but not computed): {sorted(missing)}")

    return matched, total_area_err / max(matched, 1)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    import sys

    filepath = '2176r06 Min Engagement v8.01.xlsm'

    print("=" * 60)
    print("CUTLET GEOMETRY ENGINE — Validation Run")
    print("=" * 60)

    # Step 1: Read cutter data
    print("\n1. Reading cutter data from ME file...")
    cutters, ipr, gage_radius = read_cutter_data_from_me(filepath)
    print(f"   Cutters: {len(cutters)}")
    print(f"   IPR: {ipr} in/rev")
    print(f"   Gage radius: {gage_radius}")

    # Step 2: Compute cutlets
    print("\n2. Computing cutlets...")
    results = compute_cutlets(cutters, ipr, gage_radius)
    print(f"   Cutlets computed: {len(results)}")

    # Step 3: Read ground truth
    print("\n3. Reading MassProp ground truth...")
    truth = read_massprop_ground_truth(filepath)
    print(f"   Ground truth entries: {len(truth)}")

    # Step 4: Validate
    print("\n4. Validation results:")
    matched, avg_area_err = validate_against_ground_truth(results, truth)

    # Show extra gauge cutters that were filtered by AutoCAD
    extra_names = set(r['name'] for r in results) - set(t['name'] for t in truth)
    if extra_names:
        print(f"\n   Extra computed cutters (gauge cutters skipped by AutoCAD BOUNDARY):")
        for r in results:
            if r['name'] in extra_names:
                print(f"     {r['name']:<8} Area={r['area']:.4f}")

    print(f"\n{'=' * 60}")
    if avg_area_err < 10:
        print(f"GOOD: Average area error = {avg_area_err:.1f}%")
    elif avg_area_err < 25:
        print(f"FAIR: Average area error = {avg_area_err:.1f}% — needs tuning")
    else:
        print(f"POOR: Average area error = {avg_area_err:.1f}% — fundamental issue")
    print(f"{'=' * 60}")

    return results


if __name__ == '__main__':
    main()
