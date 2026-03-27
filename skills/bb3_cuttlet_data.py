"""
Building Block 3: Cuttlet Data File

Extracts centroid X (radial), centroid Y (Z position), and cross-sectional area
from the cutlet polygons computed in BB2.

Saves an Excel file to the assembly's cuttlet_data/ folder.

Usage:
    python skills/bb3_cuttlet_data.py <source_xlsm_path> <assy_name>

Example:
    python skills/bb3_cuttlet_data.py "assemblies/Assy_2176r06/source/2176r06 Min Engagement v8.01.xlsm" Assy_2176r06
"""

import sys
import os

# Add project root to path
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, PROJECT_ROOT)

from cutlet_engine import read_cutter_data_from_me
from skills.bb2_cuttlet_plots import build_cutlet_polygons, get_blade_from_name, get_row_from_name

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def extract_cuttlet_data(cutlet_polys):
    """
    Extract centroid and area from cutlet polygons.

    Returns list of dicts sorted by centroid_x (radial position, inner to outer):
        name, blade, row, centroid_x, centroid_y, area
    """
    results = []
    for name, poly in cutlet_polys:
        if poly.is_empty or poly.area < 1e-10:
            continue
        centroid = poly.centroid
        results.append({
            'name': name,
            'blade': get_blade_from_name(name),
            'row': get_row_from_name(name),
            'centroid_x': round(-centroid.x, 4),  # negate back to positive radial
            'centroid_y': round(centroid.y, 4),
            'area': round(poly.area, 4),
        })
    results.sort(key=lambda r: r['centroid_x'])
    return results


def write_cuttlet_data_excel(data, assy_name, ipr, gage_radius, output_path):
    """Write cuttlet data to a formatted Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cuttlet Data"

    # Styles
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    num_align = Alignment(horizontal="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = f"{assy_name} — Cuttlet Data (IPR={ipr:.3f}, Gage={gage_radius:.4f})"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Headers
    headers = ["Name", "Blade", "Row", "Centroid X (in)", "Centroid Y (in)", "Area (in²)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    for i, d in enumerate(data, 4):
        values = [d['name'], d['blade'], d['row'], d['centroid_x'], d['centroid_y'], d['area']]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.alignment = num_align
            cell.border = thin_border
            if col >= 4:
                cell.number_format = '0.0000'

    # Summary row
    summary_row = len(data) + 5
    ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=summary_row, column=5, value="Count:").font = Font(bold=True)
    ws.cell(row=summary_row, column=5).alignment = Alignment(horizontal="right")
    ws.cell(row=summary_row, column=6, value=len(data)).font = Font(bold=True)

    total_area_row = summary_row + 1
    ws.cell(row=total_area_row, column=5, value="Total Area:").font = Font(bold=True)
    ws.cell(row=total_area_row, column=5).alignment = Alignment(horizontal="right")
    total_area = sum(d['area'] for d in data)
    cell = ws.cell(row=total_area_row, column=6, value=round(total_area, 4))
    cell.font = Font(bold=True)
    cell.number_format = '0.0000'

    # Column widths
    widths = [10, 8, 6, 16, 16, 14]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    wb.save(output_path)
    print(f"  Saved: {output_path}")


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    source_path = sys.argv[1]
    assy_name = sys.argv[2]

    if not os.path.exists(source_path):
        print(f"Error: Source file not found: {source_path}")
        sys.exit(1)

    # Output directory: Master/<assy_name>/
    output_dir = os.path.join(PROJECT_ROOT, "Master", assy_name)
    os.makedirs(output_dir, exist_ok=True)

    print(f"BB3: Cutlet Data — {assy_name}")
    print(f"  Source: {source_path}")

    # Read cutter data
    cutters, ipr, gage_radius = read_cutter_data_from_me(source_path)
    print(f"  Cutters: {len(cutters)}, IPR: {ipr:.3f}, Gage: {gage_radius:.4f}")

    # Build cutlet polygons (same as BB2)
    print("  Computing cutlet polygons...")
    cutlet_polys = build_cutlet_polygons(cutters, ipr, gage_radius)

    # Extract data
    data = extract_cuttlet_data(cutlet_polys)
    print(f"  Cutlets with data: {len(data)}")

    # Write Excel
    file_stem = f"{assy_name} @ {ipr:.3f} IPR"
    output_file = os.path.join(output_dir, f"{file_stem}_cutlet_data.xlsx")
    write_cuttlet_data_excel(data, assy_name, ipr, gage_radius, output_file)

    # Print summary
    print(f"\n  {'Name':<10} {'Blade':>5} {'Row':>4} {'CX (in)':>10} {'CY (in)':>10} {'Area':>10}")
    print(f"  {'-'*55}")
    for d in data:
        print(f"  {d['name']:<10} {d['blade']:>5} {d['row']:>4} {d['centroid_x']:>10.4f} {d['centroid_y']:>10.4f} {d['area']:>10.4f}")

    total = sum(d['area'] for d in data)
    print(f"  {'-'*55}")
    print(f"  {'TOTAL':<10} {'':>5} {'':>4} {'':>10} {'':>10} {total:>10.4f}")

    print(f"\nBB3 complete for {assy_name}.")


if __name__ == '__main__':
    main()
