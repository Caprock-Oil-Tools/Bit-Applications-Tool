"""
Create the Master List.xlsx file.
Shows each assembly # as rows, building block deliverables as columns.
Cells are populated with hyperlinks (named by filename) as deliverables are generated.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os

MASTER_DIR = os.path.join(os.path.dirname(__file__), "Master")
TRACKER_PATH = os.path.join(MASTER_DIR, "Master List.xlsx")

# Building block columns in progressive order
# (header_name, file_pattern) — pattern matched against filenames in each assy folder
BUILDING_BLOCKS = [
    ("Source File", "Min Engagement"),
    ("Cutlet Plot", "_cutlet_plot.png"),
    ("Cutlet Data", "_cutlet_data.xlsx"),
    ("Cutlet Forces", "_cutlet_forces.xlsx"),
    ("Bit Body Forces", "_bit_body_forces.xlsx"),
    ("Min Engagement", "_min_engagement.xlsx"),
]


def find_assemblies():
    """Find all assembly folders in the Master directory."""
    assemblies = []
    if os.path.exists(MASTER_DIR):
        for name in sorted(os.listdir(MASTER_DIR)):
            folder = os.path.join(MASTER_DIR, name)
            if os.path.isdir(folder) and name != "__pycache__":
                assemblies.append(name)
    return assemblies


def find_deliverable(assy_folder, pattern):
    """Find a file matching pattern in the assembly folder. Returns (relative_path, filename) or None."""
    folder_path = os.path.join(MASTER_DIR, assy_folder)
    if os.path.exists(folder_path):
        for f in sorted(os.listdir(folder_path)):
            if pattern in f and not f.startswith(".") and not f.startswith("~$"):
                return (os.path.join(assy_folder, f), f)
    return None


def create_tracker():
    os.makedirs(MASTER_DIR, exist_ok=True)

    # Remove old name if it exists
    old_path = os.path.join(MASTER_DIR, "master_tracker.xlsx")
    if os.path.exists(old_path):
        os.remove(old_path)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master List"

    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    link_font = Font(color="0563C1", underline="single", size=10)
    pending_font = Font(color="999999", italic=True, size=11)

    # Title row
    ws.merge_cells("A1:G1")
    ws["A1"] = "Bit Design Building Blocks \u2014 Master List"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Header row
    headers = ["Assy #"] + [bb[0] for bb in BUILDING_BLOCKS]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Set column widths
    ws.column_dimensions["A"].width = 16
    col_widths = [40, 30, 24, 24, 24, 24]
    for i, w in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i + 2)].width = w

    # Row height for header
    ws.row_dimensions[3].height = 30

    # Populate assembly rows
    assemblies = find_assemblies()
    total_linked = {bb[0]: 0 for bb in BUILDING_BLOCKS}

    for row_idx, assy_num in enumerate(assemblies, 4):
        # Assy # column
        cell = ws.cell(row=row_idx, column=1, value=assy_num)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

        # Building block columns
        for col_idx, (bb_name, bb_pattern) in enumerate(BUILDING_BLOCKS, 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = cell_align

            result = find_deliverable(assy_num, bb_pattern)
            if result:
                rel_path, filename = result
                cell.value = filename
                cell.hyperlink = rel_path
                cell.font = link_font
                total_linked[bb_name] += 1
            else:
                cell.value = "\u2014"
                cell.font = pending_font

    wb.save(TRACKER_PATH)
    print(f"Master List saved to: {TRACKER_PATH}")
    print(f"  Assemblies: {len(assemblies)}")
    for bb_name, _ in BUILDING_BLOCKS:
        print(f"    {bb_name}: {total_linked[bb_name]}/{len(assemblies)}")


if __name__ == "__main__":
    create_tracker()
