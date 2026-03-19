"""
Compute Layout Durability (0-9) and Steerability (0-9) ratings for each bit design
by analyzing cutter pocket positions and orientations from Min Engagement files.

ALL scoring inputs (back rake, side rake, positions, orientations, force directions,
redundancy, engagement patterns) are derived from the Min Engagement .xlsm files.
The main workbook is only used for bit number identification and bit size fallback.

Durability 0 = most aggressive, 9 = most durable
Steerability 0 = least steerable, 9 = most steerable

Layout classification derived from cutter geometry:
- Redundancy: detected from radial overlap between blade pairs (not from labels)
- Force direction: detected from orientation vectors (dz=axial vs dx/dy=radial)
- 6-3 offset: detected from Z-position differences between blade groups
- Blade exposure equality: detected from how uniformly blades cover the profile

IPR-dependent engagement analysis:
- Cutter naming convention: X.YZZ where X=blade, Y=row (1=primary, 2+=backup), ZZ=position
- Row 2+ cutters trail behind row 1 primaries at similar radial positions
- As IPR increases, more cutters engage; load distributions shift dynamically
- Helix angle decreases at larger radii, affecting effective rake angle
- Back-up element placement (radial offset, Z offset, degrees trailing) determines
  min engagement speed, magnitude of engagement, and rate of engagement
- Knuckles vs PDC backups have different engagement characteristics
- Some elements are intentionally underexposed (doing no work at operating IPR)
"""

import openpyxl
import os
import glob
import math
import json
import re
import warnings
import numpy as np
from collections import defaultdict

warnings.filterwarnings("ignore", category=UserWarning)

WORKBOOK_PATH = "Bit Applications Tool r2.xlsx"
BIT_DESIGNS_DIR = "Bit Designs"


def extract_cutters_from_me_file(filepath):
    """Extract cutter position and orientation data from a Min Engagement file."""
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb["Assy.Model"]

    basename = os.path.basename(filepath)
    is_v6 = "v6." in basename

    # Get bit diameter from row 1 or Settings sheet
    bit_diameter = None
    for row_idx, row in enumerate(ws.iter_rows(values_only=False), 1):
        if row_idx == 1:
            for c in row:
                if hasattr(c, "column_letter") and c.column_letter == "C" and c.value:
                    parts = str(c.value).split()
                    if parts:
                        try:
                            bit_diameter = float(parts[0])
                        except ValueError:
                            pass
            break

    if bit_diameter is None:
        try:
            ws_settings = wb["Settings"]
            for row_idx, row in enumerate(ws_settings.iter_rows(values_only=False), 1):
                if row_idx == 19:
                    for c in row:
                        if (hasattr(c, "column_letter") and c.column_letter == "C"
                                and isinstance(c.value, (int, float))):
                            bit_diameter = float(c.value)
                            break
                if bit_diameter:
                    break
        except (KeyError, Exception):
            pass

    # Extract per-cutter engagement thresholds from Settings sheet
    # Column M = min in/rev for that cutter row to engage, Column P = threshold in/rev
    engagement_thresholds = []
    try:
        ws_settings = wb["Settings"]
        for row_idx, row in enumerate(ws_settings.iter_rows(values_only=False), 1):
            if row_idx >= 14:
                m_val = None
                p_val = None
                for c in row:
                    if not hasattr(c, "column_letter"):
                        continue
                    if c.column_letter == "M" and isinstance(c.value, (int, float)):
                        m_val = c.value
                    elif c.column_letter == "P" and isinstance(c.value, (int, float)):
                        p_val = c.value
                if p_val is not None:
                    engagement_thresholds.append(p_val)
                elif m_val is not None:
                    engagement_thresholds.append(m_val)
    except (KeyError, Exception):
        pass

    cutters = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=False), 1):
        if row_idx <= 8:
            continue

        data = {}
        for c in row:
            if not hasattr(c, "column_letter") or c.value is None:
                continue
            cl = c.column_letter
            if cl == "B": data["x"] = c.value
            elif cl == "C": data["y"] = c.value
            elif cl == "D": data["z"] = c.value
            elif cl == "E": data["dx"] = c.value
            elif cl == "F": data["dy"] = c.value
            elif cl == "G": data["dz"] = c.value
            elif cl == "H": data["pocket_radius"] = c.value
            elif cl == "I": data["pocket_depth"] = c.value
            elif cl == "J": data["name"] = c.value
            elif cl == "K": data["element"] = c.value
            elif cl == "L": data["zone"] = c.value
            elif cl == "R": data["radial_pos"] = c.value
            elif cl == "Q": data["theta_deg"] = c.value
            elif cl == "AN": data["tilt_raw"] = c.value
            elif cl == "AO": data["backrake_raw"] = c.value
            elif cl == "AP": data["siderake_raw"] = c.value

        if data.get("name") is None or data.get("x") is None:
            continue

        name_str = str(data["name"])
        if "flip" in name_str.lower() or "vector" in name_str.lower():
            continue
        if "Part" in name_str:
            continue

        if "." in name_str:
            blade_str = name_str.split(".")[0]
            try:
                data["blade"] = int(blade_str)
            except ValueError:
                continue
            # Parse row number from naming convention: X.YZZ
            # Y=row (1=primary, 2+=backup), ZZ=radial position number
            suffix = name_str.split(".")[1]
            if len(suffix) >= 3:
                data["cutter_row"] = int(suffix[0])
                data["pos_num"] = int(suffix[1:])
            else:
                data["cutter_row"] = 1
                data["pos_num"] = int(suffix) if suffix.isdigit() else 0
        else:
            continue

        # Identify element type: knuckle vs PDC
        elem = data.get("element")
        elem_str = str(elem) if elem is not None else ""
        data["is_knuckle"] = "CPS" in elem_str.upper() or "knuckle" in elem_str.lower()

        # Convert back rake for v6.xx files (stored as 180 - actual_backrake)
        if is_v6 and data.get("backrake_raw") is not None:
            raw = data["backrake_raw"]
            if isinstance(raw, (int, float)) and raw > 90:
                data["backrake"] = 180.0 - raw
            elif isinstance(raw, (int, float)) and raw == 0:
                data["backrake"] = None
            else:
                data["backrake"] = raw if isinstance(raw, (int, float)) else None
        elif data.get("backrake_raw") is not None:
            raw = data["backrake_raw"]
            data["backrake"] = raw if isinstance(raw, (int, float)) else None
        else:
            data["backrake"] = None

        # Side rake conversion for v6
        if data.get("siderake_raw") is not None:
            raw = data["siderake_raw"]
            if is_v6 and isinstance(raw, (int, float)):
                data["siderake"] = abs(180.0 - raw) if raw > 90 else raw
            else:
                data["siderake"] = raw if isinstance(raw, (int, float)) else None
        else:
            data["siderake"] = None

        # Compute radial position if not available
        if data.get("radial_pos") is None:
            x, y = data.get("x"), data.get("y")
            if isinstance(x, (int, float)) and isinstance(y, (int, float)):
                data["radial_pos"] = math.sqrt(x**2 + y**2)

        cutters.append(data)

    # Deduplicate: each cutter appears twice (tip entry and base entry).
    # Keep the first occurrence per name (tip entry has the actual back rake).
    seen_names = set()
    unique_cutters = []
    for c in cutters:
        name = str(c.get("name", ""))
        if name not in seen_names:
            seen_names.add(name)
            unique_cutters.append(c)
    cutters = unique_cutters

    wb.close()
    return cutters, bit_diameter, engagement_thresholds


def detect_layout_from_geometry(cutters, gauge_radius):
    """
    Derive layout classification purely from cutter geometry:
    - Redundancy score: radial overlap between blade pairs
    - Force direction: ratio of axial vs radial orientation components
    - 6-3 offset: Z-position gap between blade groups at same radial zone
    - Blade exposure equality: how uniformly blades cover the full profile
    """
    blades_dict = defaultdict(list)
    for c in cutters:
        blades_dict[c["blade"]].append(c)

    blade_nums = sorted(blades_dict.keys())
    num_blades = len(blade_nums)

    if num_blades < 2:
        return {
            "redundancy_score": 0.0,
            "axial_force_ratio": 0.5,
            "radial_force_ratio": 0.5,
            "perpendicular_force_ratio": 0.0,
            "six_three_offset": 0.0,
            "blade_exposure_equality": 1.0,
            "effective_blade_count": num_blades,
        }

    # --- 1. REDUNDANCY: radial overlap between blade pairs ---
    # For each blade, get set of radial positions (rounded to bin)
    bin_size = gauge_radius * 0.03 if gauge_radius > 0 else 0.1  # ~3% of gauge radius
    blade_radial_bins = {}
    for b in blade_nums:
        bins = set()
        for c in blades_dict[b]:
            rp = c.get("radial_pos")
            if isinstance(rp, (int, float)):
                bins.add(round(abs(rp) / bin_size))
        blade_radial_bins[b] = bins

    # Compute pairwise radial overlap
    max_overlap = 0.0
    pair_overlaps = []
    for i, b1 in enumerate(blade_nums):
        for b2 in blade_nums[i+1:]:
            r1, r2 = blade_radial_bins[b1], blade_radial_bins[b2]
            union = len(r1 | r2)
            if union > 0:
                overlap = len(r1 & r2) / union
                pair_overlaps.append(overlap)
                max_overlap = max(max_overlap, overlap)

    # Redundancy score: average of top blade-pair overlaps
    # True redundant designs have pairs with >50% overlap
    if pair_overlaps:
        sorted_overlaps = sorted(pair_overlaps, reverse=True)
        # Take top N/2 pairs (for 6 blades, top 3 pairs represent the redundant pairs)
        top_n = max(1, num_blades // 2)
        redundancy_score = np.mean(sorted_overlaps[:top_n])
    else:
        redundancy_score = 0.0

    # --- 2. FORCE DIRECTION from orientation vectors ---
    # dz component = axial force, dx/dy = lateral/radial force
    # Perpendicular to profile = forces that push straight into the formation
    axial_sum = 0.0
    radial_sum = 0.0
    total_force = 0.0
    for c in cutters:
        dx = c.get("dx")
        dy = c.get("dy")
        dz = c.get("dz")
        if isinstance(dx, (int, float)) and isinstance(dy, (int, float)) and isinstance(dz, (int, float)):
            mag = math.sqrt(dx**2 + dy**2 + dz**2)
            if mag > 0:
                axial_sum += abs(dz) / mag
                radial_sum += math.sqrt(dx**2 + dy**2) / mag
                total_force += 1

    if total_force > 0:
        axial_force_ratio = axial_sum / total_force
        radial_force_ratio = radial_sum / total_force
    else:
        axial_force_ratio = 0.5
        radial_force_ratio = 0.5

    # Perpendicular force ratio: when forces are balanced between axial and radial
    # (redundant layouts push perpendicular to profile, which is a mix)
    perpendicular_force_ratio = 1.0 - abs(axial_force_ratio - radial_force_ratio)

    # --- 3. 6-3 OFFSET: Z-position difference between blade groups ---
    # At the nose region (35-65% of gauge radius), compare average Z per blade
    nose_lo = 0.35 * gauge_radius
    nose_hi = 0.65 * gauge_radius

    blade_z_at_nose = {}
    for b in blade_nums:
        zs = [c["z"] for c in blades_dict[b]
              if isinstance(c.get("z"), (int, float))
              and isinstance(c.get("radial_pos"), (int, float))
              and nose_lo < abs(c["radial_pos"]) < nose_hi]
        if zs:
            blade_z_at_nose[b] = np.mean(zs)

    six_three_offset = 0.0
    if len(blade_z_at_nose) >= 2:
        z_vals = list(blade_z_at_nose.values())
        z_mean = np.mean(z_vals)
        primary = [z for z in z_vals if z >= z_mean]
        secondary = [z for z in z_vals if z < z_mean]
        if primary and secondary:
            six_three_offset = abs(np.mean(primary) - np.mean(secondary))

    # --- 4. BLADE EXPOSURE EQUALITY ---
    # Measures how uniformly blades cover the radial profile.
    # For redundant designs (high pair overlap), secondary blades intentionally
    # cover less of the profile — this is a feature, not a flaw.
    # We split blades into "long" (primary) and "short" (secondary) groups
    # and measure equality within the primary group, then credit the secondary
    # group for the backup coverage they provide.
    blade_radial_ranges = {}
    for b in blade_nums:
        radii = [abs(c.get("radial_pos", 0)) for c in blades_dict[b]
                 if isinstance(c.get("radial_pos"), (int, float))]
        if radii:
            blade_radial_ranges[b] = (min(radii), max(radii))

    if blade_radial_ranges:
        coverages = [(hi - lo) for lo, hi in blade_radial_ranges.values()]
        max_coverage = max(coverages) if coverages else 1.0

        if max_coverage > 0:
            # Split blades into primary (long) and secondary (short) groups
            # Secondary blades have < 60% of max coverage
            threshold = 0.60 * max_coverage
            primary_coverages = [c for c in coverages if c >= threshold]
            secondary_coverages = [c for c in coverages if c < threshold]

            if primary_coverages and secondary_coverages:
                # Redundant design: measure equality among primary blades,
                # and give credit for having secondary blades at all
                pri_max = max(primary_coverages)
                pri_normalized = [c / pri_max for c in primary_coverages] if pri_max > 0 else [1.0]
                primary_equality = float(np.mean(pri_normalized))
                # Secondary blade presence bonus (they exist to provide backup)
                sec_bonus = min(len(secondary_coverages) / len(primary_coverages), 1.0) * 0.15
                blade_exposure_equality = min(primary_equality + sec_bonus, 1.0)
            else:
                # All blades similar coverage (F-type or SingleSet)
                normalized_coverages = [c / max_coverage for c in coverages]
                blade_exposure_equality = float(np.mean(normalized_coverages))
        else:
            blade_exposure_equality = 1.0

        # Also check how many blades start from near the center
        center_threshold = 0.15 * gauge_radius
        blades_from_center = sum(1 for lo, _ in blade_radial_ranges.values() if lo < center_threshold)
        effective_blade_count = blades_from_center if blades_from_center > 0 else num_blades
    else:
        blade_exposure_equality = 1.0
        effective_blade_count = num_blades

    return {
        "redundancy_score": float(redundancy_score),
        "axial_force_ratio": float(axial_force_ratio),
        "radial_force_ratio": float(radial_force_ratio),
        "perpendicular_force_ratio": float(perpendicular_force_ratio),
        "six_three_offset": float(six_three_offset),
        "blade_exposure_equality": float(blade_exposure_equality),
        "effective_blade_count": int(effective_blade_count),
    }


def analyze_backup_engagement(cutters, gauge_radius, engagement_thresholds):
    """
    Analyze back-up element placement and IPR-dependent engagement behavior.

    Uses the cutter naming convention (X.YZZ) to identify primary (row 1) vs
    trailing backup (row 2+) cutters, then computes:
    - Primary-backup pair offsets (radial, Z) that determine engagement behavior
    - Backup coverage: what fraction of the profile has backup elements
    - Knuckle vs PDC backup ratio
    - IPR-dependent engagement progression (how load distribution shifts with IPR)
    - Identification of non-working elements (extremely underexposed)

    Per the characterization document:
    - Radial tip offset affects min engagement speed and magnitude
    - Z tip offset affects min engagement speed (higher = engages later)
    - Degrees trailing affects sensitivity to IPR changes
    - As IPR increases, helix angle changes, effective rake changes, more cutters engage
    """
    primary_cutters = [c for c in cutters if c.get("cutter_row", 1) == 1]
    backup_cutters = [c for c in cutters if c.get("cutter_row", 1) >= 2]

    total = len(cutters)
    n_primary = len(primary_cutters)
    n_backup = len(backup_cutters)

    if total == 0:
        return _default_backup_metrics()

    backup_ratio = n_backup / total

    # --- Identify knuckle vs PDC backups ---
    knuckle_backups = [c for c in backup_cutters if c.get("is_knuckle", False)]
    pdc_backups = [c for c in backup_cutters if not c.get("is_knuckle", False)]
    knuckle_ratio = len(knuckle_backups) / n_backup if n_backup > 0 else 0.0

    # --- Match primary-backup pairs by (blade, pos_num) ---
    by_blade_pos = defaultdict(dict)
    for c in cutters:
        key = (c["blade"], c.get("pos_num", 0))
        row = c.get("cutter_row", 1)
        by_blade_pos[key][row] = c

    radial_offsets = []
    z_offsets = []
    paired_positions = 0
    unpaired_backup_positions = 0

    for key, rows in by_blade_pos.items():
        if 1 in rows and 2 in rows:
            p = rows[1]
            b = rows[2]
            rp = p.get("radial_pos")
            rb = b.get("radial_pos")
            zp = p.get("z")
            zb = b.get("z")
            if (isinstance(rp, (int, float)) and isinstance(rb, (int, float))
                    and isinstance(zp, (int, float)) and isinstance(zb, (int, float))):
                radial_offsets.append(abs(rb - rp))
                z_offsets.append(abs(zb - zp))
                paired_positions += 1
        elif 2 in rows and 1 not in rows:
            # Backup without a same-blade primary: trails behind another blade's cutter
            # Find nearest row-1 cutter across all blades at similar radial position
            b = rows[2]
            rb = b.get("radial_pos")
            if isinstance(rb, (int, float)):
                best_dist = float("inf")
                best_primary = None
                for pc in primary_cutters:
                    rp = pc.get("radial_pos")
                    if isinstance(rp, (int, float)):
                        d = abs(rp - rb)
                        if d < best_dist:
                            best_dist = d
                            best_primary = pc
                if best_primary is not None:
                    zp = best_primary.get("z")
                    zb = b.get("z")
                    if isinstance(zp, (int, float)) and isinstance(zb, (int, float)):
                        radial_offsets.append(best_dist)
                        z_offsets.append(abs(zb - zp))
                        paired_positions += 1
            unpaired_backup_positions += 1

    avg_radial_offset = float(np.mean(radial_offsets)) if radial_offsets else 0.0
    avg_z_offset = float(np.mean(z_offsets)) if z_offsets else 0.0
    max_z_offset = max(z_offsets) if z_offsets else 0.0

    # --- Backup coverage: what fraction of the radial profile has backup elements ---
    if gauge_radius > 0 and n_backup > 0:
        n_bins = 20
        bin_width = gauge_radius / n_bins
        primary_bins = set()
        backup_bins = set()
        for c in primary_cutters:
            rp = c.get("radial_pos")
            if isinstance(rp, (int, float)) and bin_width > 0:
                primary_bins.add(min(int(abs(rp) / bin_width), n_bins - 1))
        for c in backup_cutters:
            rp = c.get("radial_pos")
            if isinstance(rp, (int, float)) and bin_width > 0:
                backup_bins.add(min(int(abs(rp) / bin_width), n_bins - 1))
        backup_profile_coverage = len(backup_bins & primary_bins) / len(primary_bins) if primary_bins else 0.0
    else:
        backup_profile_coverage = 0.0

    # --- IPR-dependent engagement progression ---
    # The engagement_thresholds from Settings represent discrete IPR levels at which
    # new element size classes begin engaging. Analyze how the working cutter count
    # grows across these thresholds.
    #
    # At low IPR: only the most exposed cutters work (aggressive, fewer cutters sharing load)
    # At high IPR: backup cutters engage, redistributing load (more durable, lower ROP)
    if engagement_thresholds and len(engagement_thresholds) >= 2:
        sorted_thresholds = sorted(engagement_thresholds)
        # Engagement spread: range of IPR over which cutters progressively engage
        engagement_spread = sorted_thresholds[-1] - sorted_thresholds[0]
        # Engagement concentration: are most thresholds clustered low (aggressive)
        # or spread out (gradual engagement)?
        median_threshold = float(np.median(sorted_thresholds))
        # Low-IPR working fraction: what fraction of thresholds are below median?
        # Higher = more cutters engage early = more durable at low IPR
        low_ipr_fraction = sum(1 for t in sorted_thresholds if t <= median_threshold) / len(sorted_thresholds)
    else:
        engagement_spread = 0.25
        median_threshold = 0.25
        low_ipr_fraction = 0.5

    # --- Identify non-working elements ---
    # Extremely underexposed knuckles that are "just there" doing nothing.
    # These have very large Z offsets relative to their primaries, making them
    # essentially decorative at normal operating IPR.
    non_working_count = 0
    if z_offsets:
        # Consider a backup "non-working" if its Z offset is > 3x the average
        z_threshold = avg_z_offset * 3.0 if avg_z_offset > 0 else 0.1
        for i, zo in enumerate(z_offsets):
            if zo > z_threshold:
                non_working_count += 1
    non_working_ratio = non_working_count / n_backup if n_backup > 0 else 0.0

    # --- Helix angle effect at gauge ---
    # At larger radii, the helix angle decreases (POM becomes more circular).
    # This means gauge cutters have higher effective rake at a given IPR,
    # making them more aggressive relative to inner cutters.
    # Compute the ratio of backup cutters in the gauge region
    gauge_threshold = 0.85 * gauge_radius if gauge_radius > 0 else 4.0
    gauge_backups = [c for c in backup_cutters
                     if isinstance(c.get("radial_pos"), (int, float))
                     and abs(c["radial_pos"]) > gauge_threshold]
    gauge_backup_ratio = len(gauge_backups) / n_backup if n_backup > 0 else 0.0

    # --- Backup back rake analysis ---
    # Backup cutters typically have different (often higher) back rake than primaries
    primary_br = [c["backrake"] for c in primary_cutters
                  if c.get("backrake") is not None
                  and isinstance(c["backrake"], (int, float)) and 0 < c["backrake"] < 60]
    backup_br = [c["backrake"] for c in pdc_backups
                 if c.get("backrake") is not None
                 and isinstance(c["backrake"], (int, float)) and 0 < c["backrake"] < 60]
    avg_primary_backrake = float(np.mean(primary_br)) if primary_br else 20.0
    avg_backup_backrake = float(np.mean(backup_br)) if backup_br else 25.0
    backrake_differential = avg_backup_backrake - avg_primary_backrake

    return {
        "backup_ratio": float(backup_ratio),
        "knuckle_ratio": float(knuckle_ratio),
        "n_primary": n_primary,
        "n_backup": n_backup,
        "n_knuckle_backups": len(knuckle_backups),
        "n_pdc_backups": len(pdc_backups),
        "avg_radial_offset": float(avg_radial_offset),
        "avg_z_offset": float(avg_z_offset),
        "max_z_offset": float(max_z_offset),
        "paired_positions": paired_positions,
        "backup_profile_coverage": float(backup_profile_coverage),
        "engagement_spread": float(engagement_spread),
        "median_threshold": float(median_threshold),
        "low_ipr_fraction": float(low_ipr_fraction),
        "non_working_ratio": float(non_working_ratio),
        "gauge_backup_ratio": float(gauge_backup_ratio),
        "avg_primary_backrake": float(avg_primary_backrake),
        "avg_backup_backrake": float(avg_backup_backrake),
        "backrake_differential": float(backrake_differential),
    }


def _default_backup_metrics():
    """Return default backup metrics when no cutters are available."""
    return {
        "backup_ratio": 0.0,
        "knuckle_ratio": 0.0,
        "n_primary": 0,
        "n_backup": 0,
        "n_knuckle_backups": 0,
        "n_pdc_backups": 0,
        "avg_radial_offset": 0.0,
        "avg_z_offset": 0.0,
        "max_z_offset": 0.0,
        "paired_positions": 0,
        "backup_profile_coverage": 0.0,
        "engagement_spread": 0.25,
        "median_threshold": 0.25,
        "low_ipr_fraction": 0.5,
        "non_working_ratio": 0.0,
        "gauge_backup_ratio": 0.0,
        "avg_primary_backrake": 20.0,
        "avg_backup_backrake": 25.0,
        "backrake_differential": 5.0,
    }


def extract_engagement_summary(filepath):
    """
    Extract zone-specific engagement values directly from ME file summary area.

    The Assy.Model sheet contains a "Sceondary Cutting Structure" summary block
    with pre-computed min engagement values (ft/hr · 100 rpm) for:
      - Sec PDC (row 29): nose and taper engagement
      - Knuckle (row 30): nose and taper engagement

    Column mapping depends on ME file version:
      - v6.xx: JH = Nose engage, JI = Taper engage
      - v7.xx/v8.xx: JJ = Nose engage, JK = Taper engage

    Returns dict with keys matching compute_zone_engagement output, or empty dict
    if the summary area is not found or has no data.
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        ws = wb["Assy.Model"]
        basename = os.path.basename(filepath)
        is_v6 = "v6." in basename

        # Engagement values stored in rows 29-30 of Assy.Model
        # v6: columns JH (Nose) and JI (Taper)
        # v7/v8: columns JJ (Nose) and JK (Taper)
        nose_col = "JH" if is_v6 else "JJ"
        taper_col = "JI" if is_v6 else "JK"
        label_col = "JA" if is_v6 else "JC"

        result = {}
        for row_idx, row in enumerate(ws.iter_rows(values_only=False), 1):
            if row_idx not in (29, 30):
                continue

            # Identify row type from label
            label = None
            nose_val = None
            taper_val = None
            for c in row:
                if not hasattr(c, "column_letter"):
                    continue
                cl = c.column_letter
                if cl == label_col and isinstance(c.value, str):
                    label = c.value.strip()
                if cl == nose_col and isinstance(c.value, (int, float)):
                    nose_val = int(round(c.value))
                if cl == taper_col and isinstance(c.value, (int, float)):
                    taper_val = int(round(c.value))

            if label is None:
                # Try the alternate label column (some files use JA, some JC)
                for c in row:
                    if not hasattr(c, "column_letter"):
                        continue
                    cl = c.column_letter
                    alt_col = "JC" if is_v6 else "JA"
                    if cl == alt_col and isinstance(c.value, str):
                        label = c.value.strip()

            if label and "sec" in label.lower() and "pdc" in label.lower():
                if nose_val is not None and nose_val > 0:
                    result["nose_pdc_engage"] = nose_val
                if taper_val is not None and taper_val > 0:
                    result["taper_pdc_engage"] = taper_val
            elif label and "knuckle" in label.lower():
                if nose_val is not None and nose_val > 0:
                    result["nose_knuckle_engage"] = nose_val
                if taper_val is not None and taper_val > 0:
                    result["taper_knuckle_engage"] = taper_val

        wb.close()
        return result
    except Exception:
        return {}


def compute_zone_engagement(cutters, gauge_radius, engagement_thresholds):
    """
    Compute zone-specific minimum engagement values for BAT columns BA-BE.

    Returns dict with keys:
      six_three_engage: min ft/hr@100rpm for pri PDCs on sec blades to engage (BA)
      nose_pdc_engage:  min ft/hr@100rpm for sec PDC backups at nose (BB)
      taper_pdc_engage: min ft/hr@100rpm for sec PDC backups at taper (BC)
      nose_knuckle_engage: min ft/hr@100rpm for knuckle backups at nose (BD)
      taper_knuckle_engage: min ft/hr@100rpm for knuckle backups at taper (BE)

    Zone boundaries (fraction of gauge radius):
      Nose:  0.30 - 0.60
      Taper: 0.60 - 0.85
    """
    if not cutters or gauge_radius <= 0:
        return {}

    primary_cutters = [c for c in cutters if c.get("cutter_row", 1) == 1]
    backup_cutters = [c for c in cutters if c.get("cutter_row", 1) >= 2]

    if not backup_cutters:
        return {}

    # Build sorted IPR thresholds from Settings (in/rev values)
    sorted_thresholds = sorted(engagement_thresholds) if engagement_thresholds else []

    # For each backup cutter, find its paired primary and compute Z offset.
    # Then estimate the IPR at which it engages from the Settings thresholds.
    # The backup with larger Z offset engages at higher IPR (needs deeper cut).
    # Map Z offset to closest engagement threshold class.

    # First, collect all backup Z offsets to determine the distribution
    backup_z_offsets = []
    for bc in backup_cutters:
        rb = bc.get("radial_pos")
        zb = bc.get("z")
        if not (isinstance(rb, (int, float)) and isinstance(zb, (int, float))):
            continue
        # Find nearest primary cutter
        best_dist = float("inf")
        best_primary = None
        for pc in primary_cutters:
            rp = pc.get("radial_pos")
            if isinstance(rp, (int, float)):
                d = abs(rp - rb)
                if d < best_dist:
                    best_dist = d
                    best_primary = pc
        if best_primary is not None:
            zp = best_primary.get("z")
            if isinstance(zp, (int, float)):
                backup_z_offsets.append(abs(zb - zp))

    if not backup_z_offsets:
        return {}

    # Sort z offsets and map to engagement threshold classes
    # The ME file's threshold classes correspond to increasing engagement difficulty
    # Map each backup's z_offset rank to a threshold class proportionally
    all_z_sorted = sorted(set(backup_z_offsets))

    def z_offset_to_fthr(z_off):
        """Convert a backup cutter's Z offset to estimated ft/hr at 100 RPM."""
        if not sorted_thresholds:
            # Fallback: use linear estimate. Z offset in inches → in/rev ≈ z_off * 2
            # (rough geometric approximation for helical engagement)
            ipr_est = z_off * 2.0
            return round(ipr_est * 500)  # Convert in/rev to ft/hr at 100 RPM

        # Map z_offset rank to threshold class
        if len(all_z_sorted) <= 1:
            idx = len(sorted_thresholds) // 2
        else:
            rank = all_z_sorted.index(z_off) if z_off in all_z_sorted else 0
            frac = rank / (len(all_z_sorted) - 1)
            idx = int(frac * (len(sorted_thresholds) - 1))
            idx = min(idx, len(sorted_thresholds) - 1)

        ipr = sorted_thresholds[idx]
        return round(ipr * 500)  # in/rev → ft/hr at 100 RPM

    # Classify backup cutters by zone and type, compute engagement values
    nose_lo = 0.30 * gauge_radius
    nose_hi = 0.60 * gauge_radius
    taper_lo = 0.60 * gauge_radius
    taper_hi = 0.85 * gauge_radius

    zone_engage = {
        "nose_pdc": [], "taper_pdc": [],
        "nose_knuckle": [], "taper_knuckle": [],
    }

    # Also track 6-3 engagement: primary PDCs on secondary (odd-numbered) blades
    num_blades = len(set(c["blade"] for c in cutters if "blade" in c))
    # In a 6-3 layout, secondary blades are the higher-numbered blades
    # (e.g., blades 4-6 in a 6-blade, or blades 7-9 in a 9-blade design)
    # Secondary blade primaries engage later due to Z offset from primary blades
    pri_blade_cutters = defaultdict(list)
    for c in primary_cutters:
        rp = c.get("radial_pos")
        z = c.get("z")
        if isinstance(rp, (int, float)) and isinstance(z, (int, float)):
            pri_blade_cutters[c["blade"]].append(c)

    six_three_engage_vals = []
    if num_blades >= 6:
        blade_nums = sorted(pri_blade_cutters.keys())
        if len(blade_nums) >= 6:
            # Primary blades = first half, secondary blades = second half
            mid = len(blade_nums) // 2
            pri_blades = set(blade_nums[:mid])
            sec_blades = set(blade_nums[mid:])

            # For each primary cutter on a secondary blade, find nearest
            # primary cutter on a primary blade and compute Z offset
            for sb in sec_blades:
                for sc in pri_blade_cutters[sb]:
                    rs = sc.get("radial_pos")
                    zs = sc.get("z")
                    if not (isinstance(rs, (int, float)) and isinstance(zs, (int, float))):
                        continue
                    best_dist = float("inf")
                    best_z = None
                    for pb in pri_blades:
                        for pc in pri_blade_cutters[pb]:
                            rp = pc.get("radial_pos")
                            if isinstance(rp, (int, float)):
                                d = abs(rp - rs)
                                if d < best_dist:
                                    best_dist = d
                                    best_z = pc.get("z")
                    if best_z is not None and isinstance(best_z, (int, float)):
                        z_off = abs(zs - best_z)
                        six_three_engage_vals.append(z_offset_to_fthr(z_off))

    bi = 0
    for bc in backup_cutters:
        rb = bc.get("radial_pos")
        zb = bc.get("z")
        is_knuckle = bc.get("is_knuckle", False)
        if not (isinstance(rb, (int, float)) and isinstance(zb, (int, float))):
            continue

        r_abs = abs(rb)
        # Find paired primary Z offset
        best_dist = float("inf")
        best_primary = None
        for pc in primary_cutters:
            rp = pc.get("radial_pos")
            if isinstance(rp, (int, float)):
                d = abs(rp - rb)
                if d < best_dist:
                    best_dist = d
                    best_primary = pc
        if best_primary is None:
            continue
        zp = best_primary.get("z")
        if not isinstance(zp, (int, float)):
            continue

        z_off = abs(zb - zp)
        fthr = z_offset_to_fthr(z_off)

        # Classify by zone
        if nose_lo <= r_abs < nose_hi:
            zone = "nose"
        elif taper_lo <= r_abs < taper_hi:
            zone = "taper"
        else:
            continue  # cone or gauge - not in BA-BE columns

        typ = "knuckle" if is_knuckle else "pdc"
        zone_engage[f"{zone}_{typ}"].append(fthr)

    result = {}
    if six_three_engage_vals:
        result["six_three_engage"] = min(six_three_engage_vals)
    if zone_engage["nose_pdc"]:
        result["nose_pdc_engage"] = min(zone_engage["nose_pdc"])
    if zone_engage["taper_pdc"]:
        result["taper_pdc_engage"] = min(zone_engage["taper_pdc"])
    if zone_engage["nose_knuckle"]:
        result["nose_knuckle_engage"] = min(zone_engage["nose_knuckle"])
    if zone_engage["taper_knuckle"]:
        result["taper_knuckle_engage"] = min(zone_engage["taper_knuckle"])

    return result


def compute_metrics(cutters, bit_diameter, engagement_thresholds):
    """
    Compute durability and steerability metrics purely from Min Engagement data.
    No inputs from the main workbook except bit_diameter as fallback.
    """
    if not cutters:
        return None

    blades = defaultdict(list)
    for c in cutters:
        blades[c["blade"]].append(c)

    num_blades = len(blades)
    if num_blades == 0:
        return None

    bit_radius = bit_diameter / 2.0 if bit_diameter else None

    # Gauge radius from cutter data
    all_radii = [abs(c.get("radial_pos", 0)) for c in cutters
                 if isinstance(c.get("radial_pos"), (int, float))]
    gauge_radius = bit_radius if bit_radius else (max(all_radii) if all_radii else 4.0)

    total_cutters = len(cutters)
    avg_cutters_per_blade = total_cutters / num_blades

    # --- LAYOUT DETECTION FROM GEOMETRY ---
    layout = detect_layout_from_geometry(cutters, gauge_radius)

    # --- BACKUP ENGAGEMENT ANALYSIS ---
    backup = analyze_backup_engagement(cutters, gauge_radius, engagement_thresholds)

    # --- ZONE-SPECIFIC ENGAGEMENT (for BAT columns BA-BE) ---
    zone_engagement = compute_zone_engagement(cutters, gauge_radius, engagement_thresholds)

    # --- BACK RAKE STATISTICS (from ME file) ---
    all_backrakes = [c["backrake"] for c in cutters
                     if c.get("backrake") is not None
                     and isinstance(c["backrake"], (int, float)) and 0 < c["backrake"] < 60]
    avg_backrake = float(np.mean(all_backrakes)) if all_backrakes else 20.0

    # Back rake by zone (cone <30%, nose 30-60%, taper 60-85%, gauge >85%)
    cone_br = [c["backrake"] for c in cutters
               if c.get("backrake") and isinstance(c["backrake"], (int, float))
               and 0 < c["backrake"] < 60
               and isinstance(c.get("radial_pos"), (int, float))
               and abs(c["radial_pos"]) < 0.30 * gauge_radius]
    nose_br = [c["backrake"] for c in cutters
               if c.get("backrake") and isinstance(c["backrake"], (int, float))
               and 0 < c["backrake"] < 60
               and isinstance(c.get("radial_pos"), (int, float))
               and 0.30 * gauge_radius <= abs(c["radial_pos"]) < 0.60 * gauge_radius]
    taper_br = [c["backrake"] for c in cutters
                if c.get("backrake") and isinstance(c["backrake"], (int, float))
                and 0 < c["backrake"] < 60
                and isinstance(c.get("radial_pos"), (int, float))
                and 0.60 * gauge_radius <= abs(c["radial_pos"]) < 0.85 * gauge_radius]
    gauge_br = [c["backrake"] for c in cutters
                if c.get("backrake") and isinstance(c["backrake"], (int, float))
                and 0 < c["backrake"] < 60
                and isinstance(c.get("radial_pos"), (int, float))
                and abs(c["radial_pos"]) >= 0.85 * gauge_radius]

    avg_cone_backrake = float(np.mean(cone_br)) if cone_br else 15.0
    avg_nose_backrake = float(np.mean(nose_br)) if nose_br else 15.0
    avg_gauge_backrake = float(np.mean(gauge_br)) if gauge_br else 30.0

    # --- CUTTER DENSITY AND SPACING (from ME file) ---
    # Size-adjusted: normalize by cutter diameter so that many small cutters
    # count proportionally to fewer large cutters covering the same area
    raw_density = total_cutters / gauge_radius if gauge_radius > 0 else total_cutters / 4.0
    # Compute average primary cutter diameter for size adjustment
    _pri_radii = [c.get("pocket_radius") for c in cutters
                  if isinstance(c.get("pocket_radius"), (int, float))
                  and c.get("pocket_radius", 0) > 0.1
                  and "." in str(c.get("name", ""))
                  and str(c.get("name", "")).split(".")[1][:1] == "1"]
    _avg_pri_dia_mm = (sum(_pri_radii) / len(_pri_radii) * 2 * 25.4) if _pri_radii else 16.0
    # Scale density: 16mm is reference; smaller cutters get proportional boost
    cutter_density = raw_density * (16.0 / _avg_pri_dia_mm)

    # Radial zone coverage: how many blades present per zone
    num_zones = 10
    zone_width = gauge_radius / num_zones if gauge_radius > 0 else 0.5
    zone_blade_coverage = defaultdict(set)
    for c in cutters:
        rp = c.get("radial_pos")
        if isinstance(rp, (int, float)) and zone_width > 0:
            zone_idx = min(int(abs(rp) / zone_width), num_zones - 1)
            zone_blade_coverage[zone_idx].add(c["blade"])

    avg_blades_per_zone = float(np.mean([len(v) for v in zone_blade_coverage.values()])) if zone_blade_coverage else 1.0

    # Spacing uniformity per blade
    spacing_cvs = []
    for blade_num, blade_cutters in blades.items():
        radii = sorted([abs(c.get("radial_pos", 0)) for c in blade_cutters
                        if isinstance(c.get("radial_pos"), (int, float))])
        if len(radii) > 1:
            spacings = [radii[i+1] - radii[i] for i in range(len(radii)-1)]
            if spacings:
                mean_sp = np.mean(spacings)
                if mean_sp > 0:
                    spacing_cvs.append(float(np.std(spacings) / mean_sp))

    avg_spacing_cv = float(np.mean(spacing_cvs)) if spacing_cvs else 0.5

    # --- GAUGE REGION ANALYSIS (from ME file) ---
    gauge_threshold = 0.85 * gauge_radius
    gauge_cutters = [c for c in cutters
                     if isinstance(c.get("radial_pos"), (int, float))
                     and abs(c["radial_pos"]) > gauge_threshold]
    gauge_cutter_ratio = len(gauge_cutters) / total_cutters if total_cutters > 0 else 0

    # --- SIDE RAKE (from ME file) ---
    all_siderakes = [c["siderake"] for c in cutters
                     if c.get("siderake") is not None
                     and isinstance(c["siderake"], (int, float)) and abs(c["siderake"]) < 30]
    avg_siderake = float(np.mean([abs(s) for s in all_siderakes])) if all_siderakes else 0.0

    # --- FORCE IMBALANCE (from ME file orientation vectors) ---
    fx_sum, fy_sum, force_count = 0.0, 0.0, 0
    for c in cutters:
        dx, dy = c.get("dx"), c.get("dy")
        if isinstance(dx, (int, float)) and isinstance(dy, (int, float)):
            fx_sum += dx
            fy_sum += dy
            force_count += 1
    lateral_resultant = math.sqrt(fx_sum**2 + fy_sum**2) / force_count if force_count > 0 else 0.0

    # --- PROFILE DEPTH (from ME file Z positions) ---
    all_z = [c["z"] for c in cutters if isinstance(c.get("z"), (int, float))]
    z_range = (max(all_z) - min(all_z)) if len(all_z) > 1 else 0.0

    # --- ENGAGEMENT THRESHOLDS (from ME Settings sheet) ---
    # Higher min engagement = more aggressive (secondary cutters don't engage easily)
    avg_engagement_threshold = float(np.mean(engagement_thresholds)) if engagement_thresholds else 0.25
    max_engagement_threshold = max(engagement_thresholds) if engagement_thresholds else 0.5

    # --- POCKET DEPTH (from ME file) ---
    depths = [c["pocket_depth"] for c in cutters
              if isinstance(c.get("pocket_depth"), (int, float))]
    avg_depth = float(np.mean(depths)) if depths else 0.4

    # --- PRIMARY CUTTER DIAMETER (from ME file pocket_radius) ---
    # Primary cutters have names matching X.1YY pattern (second digit after . is '1')
    primary_radii = []
    for c in cutters:
        r = c.get("pocket_radius")
        name = str(c.get("name", ""))
        if isinstance(r, (int, float)) and r > 0.1 and "." in name:
            parts = name.split(".")
            if len(parts) == 2 and len(parts[1]) >= 1 and parts[1][0] == "1":
                primary_radii.append(r)
    avg_primary_cutter_dia_mm = (
        round(sum(primary_radii) / len(primary_radii) * 2 * 25.4, 1)
        if primary_radii else 16.0  # default to 16mm if unknown
    )

    return {
        # Layout geometry (all from ME file)
        "redundancy_score": layout["redundancy_score"],
        "axial_force_ratio": layout["axial_force_ratio"],
        "radial_force_ratio": layout["radial_force_ratio"],
        "perpendicular_force_ratio": layout["perpendicular_force_ratio"],
        "six_three_offset": layout["six_three_offset"],
        "blade_exposure_equality": layout["blade_exposure_equality"],
        "effective_blade_count": layout["effective_blade_count"],
        # Cutter geometry (all from ME file)
        "total_cutters": total_cutters,
        "num_blades": num_blades,
        "avg_cutters_per_blade": avg_cutters_per_blade,
        "avg_backrake": avg_backrake,
        "avg_cone_backrake": avg_cone_backrake,
        "avg_nose_backrake": avg_nose_backrake,
        "avg_gauge_backrake": avg_gauge_backrake,
        "cutter_density": cutter_density,
        "avg_blades_per_zone": avg_blades_per_zone,
        "avg_spacing_cv": avg_spacing_cv,
        "gauge_cutter_ratio": gauge_cutter_ratio,
        "avg_siderake": avg_siderake,
        "lateral_resultant": lateral_resultant,
        "z_range": z_range,
        "avg_engagement_threshold": avg_engagement_threshold,
        "max_engagement_threshold": max_engagement_threshold,
        "avg_depth": avg_depth,
        "gauge_radius": gauge_radius,
        "avg_primary_cutter_dia_mm": avg_primary_cutter_dia_mm,
        # Backup engagement analysis (IPR-dependent behavior)
        "backup_ratio": backup["backup_ratio"],
        "knuckle_ratio": backup["knuckle_ratio"],
        "n_primary": backup["n_primary"],
        "n_backup": backup["n_backup"],
        "n_knuckle_backups": backup["n_knuckle_backups"],
        "n_pdc_backups": backup["n_pdc_backups"],
        "avg_radial_offset": backup["avg_radial_offset"],
        "avg_z_offset": backup["avg_z_offset"],
        "max_z_offset": backup["max_z_offset"],
        "paired_positions": backup["paired_positions"],
        "backup_profile_coverage": backup["backup_profile_coverage"],
        "engagement_spread": backup["engagement_spread"],
        "median_threshold": backup["median_threshold"],
        "low_ipr_fraction": backup["low_ipr_fraction"],
        "non_working_ratio": backup["non_working_ratio"],
        "gauge_backup_ratio": backup["gauge_backup_ratio"],
        "avg_primary_backrake": backup["avg_primary_backrake"],
        "avg_backup_backrake": backup["avg_backup_backrake"],
        "backrake_differential": backup["backrake_differential"],
        # Zone-specific engagement (for BAT columns BA-BE)
        "zone_engagement": zone_engagement,
    }


def compute_cutlet_metrics(cutlet_results, cutters_from_me, gage_radius):
    """
    Compute durability and steerability metrics directly from cutlet geometry.

    This uses DIRECT measurements of what each cutter does:
    cutlet area = rock removed per revolution, perimeter = chamfer contact length.

    Key physics captured:
    - Force distribution SHAPE (normalized areas, not magnitudes)
    - Chamfer fraction: 0.020" x 45° chamfer on all PDCs. Smaller cutlets have
      proportionally more chamfer area → less efficient cutting but more durable.
    - Constant volume principle: for a given hole size and IPR, total volume removed
      is constant regardless of cutter count.

    Inputs:
        cutlet_results: list of dicts from cutlet_engine.compute_cutlets()
            Each has: name, centroid_x, centroid_y, area, perimeter
        cutters_from_me: list of dicts from extract_cutters_from_me_file()
            Each has: blade, cutter_row, backrake, radial_pos, z, etc.
        gage_radius: float, bit radius in inches

    Returns dict of cutlet-derived metrics.
    """
    if not cutlet_results:
        return None

    CHAMFER_WIDTH = 0.020  # inches, 45° chamfer on all PDCs

    areas = [r['area'] for r in cutlet_results]
    perimeters = [r.get('perimeter', 0) for r in cutlet_results]
    centroids_x = [r['centroid_x'] for r in cutlet_results]
    centroids_y = [r['centroid_y'] for r in cutlet_results]

    # Parse blade/row from cutter names
    pri_areas = []
    bkp_areas = []
    cone_areas = []      # < 30% of gage
    nose_areas = []      # 30-60%
    shoulder_areas = []  # 60-85%
    gauge_areas = []     # > 85%

    blade_areas = defaultdict(list)  # areas grouped by blade

    # Per-cutlet chamfer fractions
    chamfer_fractions = []

    # Normalized force distribution: area fractions across radial bins
    n_radial_bins = 10
    radial_bin_width = gage_radius / n_radial_bins if gage_radius > 0 else 0.5
    radial_bin_areas = [0.0] * n_radial_bins

    for r in cutlet_results:
        name_str = str(r['name'])
        parts = name_str.split('.')
        blade = int(parts[0]) if len(parts) == 2 else 1
        suffix = parts[1] if len(parts) == 2 else '100'
        cutter_row = int(suffix[0]) if len(suffix) >= 1 else 1

        blade_areas[blade].append(r['area'])

        # Zone classification by centroid radial position
        radial_frac = r['centroid_x'] / gage_radius if gage_radius > 0 else 0
        if radial_frac < 0.30:
            cone_areas.append(r['area'])
        elif radial_frac < 0.60:
            nose_areas.append(r['area'])
        elif radial_frac < 0.85:
            shoulder_areas.append(r['area'])
        else:
            gauge_areas.append(r['area'])

        if cutter_row == 1:
            pri_areas.append(r['area'])
        else:
            bkp_areas.append(r['area'])

        # Radial bin for force distribution
        bin_idx = min(int(r['centroid_x'] / radial_bin_width), n_radial_bins - 1) if radial_bin_width > 0 else 0
        radial_bin_areas[bin_idx] += r['area']

        # Chamfer fraction for this cutlet
        perim = r.get('perimeter', 0)
        if r['area'] > 0 and perim > 0:
            # Chamfer area ≈ perimeter × chamfer_width (for small chamfer relative to cutlet)
            chamfer_area = perim * CHAMFER_WIDTH
            cf = min(chamfer_area / r['area'], 1.0)  # cap at 1.0
            chamfer_fractions.append(cf)
        else:
            chamfer_fractions.append(0.0)

    n_cutlets = len(areas)
    total_area = sum(areas)
    mean_area = np.mean(areas)
    area_cv = float(np.std(areas) / mean_area) if mean_area > 0 else 0
    max_area = max(areas)
    min_area = min(areas)
    max_min_ratio = max_area / min_area if min_area > 0 else 999

    # --- CHAMFER METRICS ---
    avg_chamfer_fraction = float(np.mean(chamfer_fractions)) if chamfer_fractions else 0.0
    # Area-weighted chamfer fraction (bigger cutlets contribute more)
    weighted_chamfer = float(sum(cf * a for cf, a in zip(chamfer_fractions, areas)) / total_area) if total_area > 0 else 0.0

    # --- FORCE DISTRIBUTION SHAPE (normalized — NOT magnitudes) ---
    # Normalize bin areas to probability distribution
    total_bin_area = sum(radial_bin_areas)
    if total_bin_area > 0:
        force_dist = [a / total_bin_area for a in radial_bin_areas]
    else:
        force_dist = [1.0 / n_radial_bins] * n_radial_bins

    # Force distribution uniformity: entropy-based (max entropy = perfectly uniform)
    # Shannon entropy of the normalized distribution
    force_entropy = 0.0
    for p in force_dist:
        if p > 0:
            force_entropy -= p * math.log(p)
    max_entropy = math.log(n_radial_bins)  # perfectly uniform distribution
    force_uniformity = force_entropy / max_entropy if max_entropy > 0 else 0

    # Force concentration: what fraction of total force is in the most loaded bin?
    max_bin_frac = max(force_dist) if force_dist else 0
    # Force spread: how many bins have >5% of total force?
    active_bins = sum(1 for p in force_dist if p > 0.05)
    force_spread = active_bins / n_radial_bins

    # --- LOAD BALANCE: how uniformly is cutting work distributed? ---
    # Gini coefficient: 0 = perfect equality, 1 = all work on one cutter
    sorted_areas = sorted(areas)
    n = len(sorted_areas)
    gini = float((2 * np.sum((np.arange(1, n + 1) * sorted_areas)) - (n + 1) * total_area)
                 / (n * total_area)) if total_area > 0 else 0

    # --- BLADE LOAD BALANCE ---
    blade_totals = [sum(v) for v in blade_areas.values()]
    blade_cv = float(np.std(blade_totals) / np.mean(blade_totals)) if blade_totals and np.mean(blade_totals) > 0 else 0

    # --- WORK PER CUTTER ---
    work_per_cutter = total_area / n_cutlets if n_cutlets > 0 else 0

    # --- BACKUP ENGAGEMENT ---
    n_backup_cutlets = len(bkp_areas)
    backup_cutlet_fraction = n_backup_cutlets / n_cutlets if n_cutlets > 0 else 0
    backup_area_fraction = sum(bkp_areas) / total_area if total_area > 0 and bkp_areas else 0

    # --- GAUGE REGION ---
    gauge_area_total = sum(gauge_areas)
    gauge_area_fraction = gauge_area_total / total_area if total_area > 0 else 0
    n_gauge_cutlets = len(gauge_areas)

    # --- ZONE DISTRIBUTION (normalized — force distribution shape) ---
    cone_frac = sum(cone_areas) / total_area if total_area > 0 else 0
    nose_frac = sum(nose_areas) / total_area if total_area > 0 else 0
    shoulder_frac = sum(shoulder_areas) / total_area if total_area > 0 else 0
    gauge_frac = gauge_area_fraction

    # --- BACKRAKE from ME file (cutlets don't encode orientation) ---
    all_br = [c["backrake"] for c in cutters_from_me
              if c.get("backrake") is not None
              and isinstance(c["backrake"], (int, float)) and 0 < c["backrake"] < 60]
    avg_backrake = float(np.mean(all_br)) if all_br else 20.0

    gauge_br = [c["backrake"] for c in cutters_from_me
                if c.get("backrake") and isinstance(c["backrake"], (int, float))
                and 0 < c["backrake"] < 60
                and isinstance(c.get("radial_pos"), (int, float))
                and abs(c["radial_pos"]) >= 0.85 * gage_radius]
    avg_gauge_backrake = float(np.mean(gauge_br)) if gauge_br else 30.0

    cone_br = [c["backrake"] for c in cutters_from_me
               if c.get("backrake") and isinstance(c["backrake"], (int, float))
               and 0 < c["backrake"] < 60
               and isinstance(c.get("radial_pos"), (int, float))
               and abs(c["radial_pos"]) < 0.30 * gage_radius]
    avg_cone_backrake = float(np.mean(cone_br)) if cone_br else 15.0

    # --- SIDERAKE ---
    all_sr = [abs(c["siderake"]) for c in cutters_from_me
              if c.get("siderake") is not None
              and isinstance(c["siderake"], (int, float)) and abs(c["siderake"]) < 30]
    avg_siderake = float(np.mean(all_sr)) if all_sr else 0.0

    # --- PROFILE DEPTH from cutlet centroids ---
    y_range = max(centroids_y) - min(centroids_y) if len(centroids_y) > 1 else 0

    # --- CUTTER SIZE from ME file ---
    primary_radii = []
    for c in cutters_from_me:
        r = c.get("pocket_radius")
        name = str(c.get("name", ""))
        if isinstance(r, (int, float)) and r > 0.1 and "." in name:
            parts = name.split(".")
            if len(parts) == 2 and len(parts[1]) >= 1 and parts[1][0] == "1":
                primary_radii.append(r)
    avg_primary_dia_mm = (
        round(sum(primary_radii) / len(primary_radii) * 2 * 25.4, 1)
        if primary_radii else 16.0
    )

    # --- BLADE COUNT ---
    num_blades = len(blade_areas)

    # --- BIT SIZE ---
    bit_diameter = gage_radius * 2

    # --- CUTTER DENSITY: cutlets per sq inch of bit face ---
    # More cutlets sharing the same total volume = smaller individual cutlets
    # = less force per cutter = more durable (constant volume principle)
    bit_face_area = math.pi * gage_radius ** 2
    cutter_density = n_cutlets / bit_face_area if bit_face_area > 0 else 0

    # --- MAX-TO-MEAN RATIO: peak cutter overload indicator ---
    # If the largest cutlet is much bigger than average, one cutter is doing
    # disproportionate work and will wear/break first
    max_mean_ratio = max_area / mean_area if mean_area > 0 else 1.0

    # --- PRIMARY vs SECONDARY BLADE ROW 1 CUTLET SIZE ANALYSIS ---
    # Primary blades start near center; secondary blades start mid-radius.
    # Compare row 1 cutlet sizes at overlapping radial positions.
    # Large differential = primary blades doing most of the work = more aggressive.
    blade_min_radial = {}
    blade_row1_cutlets = defaultdict(list)  # blade -> [(radial, area)]
    for r in cutlet_results:
        name_str = str(r['name'])
        parts = name_str.split('.')
        if len(parts) != 2:
            continue
        try:
            blade_num = int(parts[0])
        except ValueError:
            continue
        suffix = parts[1]
        cutter_row = int(suffix[0]) if len(suffix) >= 1 and suffix[0].isdigit() else 1

        radial = r['centroid_x']
        if blade_num not in blade_min_radial or radial < blade_min_radial[blade_num]:
            blade_min_radial[blade_num] = radial
        if cutter_row == 1:
            blade_row1_cutlets[blade_num].append((radial, r['area']))

    # Classify blades: primary start near center, secondary start mid-radius
    primary_blades = set()
    secondary_blades = set()
    center_threshold = 0.25 * gage_radius
    for bnum, min_r in blade_min_radial.items():
        if min_r < center_threshold:
            primary_blades.add(bnum)
        else:
            secondary_blades.add(bnum)

    # Compare row 1 cutlet areas at overlapping radial positions
    pri_sec_area_ratio = 1.0  # default: balanced
    n_primary_blades = len(primary_blades)
    n_secondary_blades = len(secondary_blades)

    if primary_blades and secondary_blades:
        # Find the radial range where secondary blades have cutters
        sec_radials = []
        for b in secondary_blades:
            if b in blade_row1_cutlets:
                sec_radials.extend([r for r, _ in blade_row1_cutlets[b]])
        if sec_radials:
            sec_min_r = min(sec_radials)
            sec_max_r = max(sec_radials)

            # Collect primary blade row 1 areas in the secondary blade radial range
            pri_areas_overlap = []
            for b in primary_blades:
                if b in blade_row1_cutlets:
                    for r, a in blade_row1_cutlets[b]:
                        if sec_min_r * 0.9 <= r <= sec_max_r * 1.1:
                            pri_areas_overlap.append(a)

            sec_areas_all = []
            for b in secondary_blades:
                if b in blade_row1_cutlets:
                    sec_areas_all.extend([a for _, a in blade_row1_cutlets[b]])

            if pri_areas_overlap and sec_areas_all:
                pri_mean = float(np.mean(pri_areas_overlap))
                sec_mean = float(np.mean(sec_areas_all))
                if sec_mean > 0:
                    pri_sec_area_ratio = pri_mean / sec_mean

    return {
        # Cutlet-derived
        "n_cutlets": n_cutlets,
        "total_area": float(total_area),
        "mean_area": float(mean_area),
        "area_cv": float(area_cv),
        "max_area": float(max_area),
        "min_area": float(min_area),
        "max_min_ratio": float(max_min_ratio),
        "gini": float(gini),
        "blade_cv": float(blade_cv),
        "work_per_cutter": float(work_per_cutter),
        "backup_cutlet_fraction": float(backup_cutlet_fraction),
        "backup_area_fraction": float(backup_area_fraction),
        "gauge_area_total": float(gauge_area_total),
        "gauge_area_fraction": float(gauge_area_fraction),
        "n_gauge_cutlets": n_gauge_cutlets,
        "cone_frac": float(cone_frac),
        "nose_frac": float(nose_frac),
        "shoulder_frac": float(shoulder_frac),
        "gauge_frac": float(gauge_frac),
        "y_range": float(y_range),
        # Chamfer-derived
        "avg_chamfer_fraction": float(avg_chamfer_fraction),
        "weighted_chamfer_fraction": float(weighted_chamfer),
        # Force distribution shape
        "force_uniformity": float(force_uniformity),
        "force_spread": float(force_spread),
        "max_bin_frac": float(max_bin_frac),
        "force_dist": [round(f, 4) for f in force_dist],
        # Orientation-derived (from ME file)
        "avg_backrake": float(avg_backrake),
        "avg_gauge_backrake": float(avg_gauge_backrake),
        "avg_cone_backrake": float(avg_cone_backrake),
        "avg_siderake": float(avg_siderake),
        "avg_primary_dia_mm": float(avg_primary_dia_mm),
        "num_blades": num_blades,
        "gage_radius": float(gage_radius),
        "bit_diameter": float(bit_diameter),
        # Layout structure metrics
        "cutter_density": float(cutter_density),
        "max_mean_ratio": float(max_mean_ratio),
        "pri_sec_area_ratio": float(pri_sec_area_ratio),
        "n_primary_blades": n_primary_blades,
        "n_secondary_blades": n_secondary_blades,
    }


def main():
    from cutlet_engine import compute_cutlets

    print("Loading main workbook (for bit numbers and size fallback only)...")
    wb_main = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    ws = wb_main["Sheet1"]

    bits = []
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=False):
        bit_data = {}
        for c in row:
            cl = c.column_letter
            if cl == "B": bit_data["bit_num"] = c.value
            elif cl == "X": bit_data["size"] = c.value
            elif cl == "AR": bit_data["layout_type"] = c.value  # for display only
        if bit_data.get("bit_num") is not None:
            bit_data["row_num"] = row[0].row
            bits.append(bit_data)

    print(f"Found {len(bits)} bits in workbook")

    # Find and match Min Engagement files
    all_me_files = glob.glob(f"{BIT_DESIGNS_DIR}/**/*Min Engagement*.xlsm", recursive=True)
    print(f"Found {len(all_me_files)} Min Engagement files")

    bit_me_map = {}
    for f in all_me_files:
        folder_parts = f.replace("\\", "/").split("/")
        folder_name = folder_parts[1] if len(folder_parts) > 1 else ""
        for bit in bits:
            bn = str(int(bit["bit_num"])) if isinstance(bit["bit_num"], (int, float)) else str(bit["bit_num"])
            if folder_name == bn or bn in os.path.basename(f).split(" ")[0].replace(".", "").replace("r", ""):
                version = 0
                bname = os.path.basename(f)
                if "v8." in bname: version = 8
                elif "v7." in bname: version = 7
                elif "v6." in bname: version = 6
                is_variant = any(x in bname.upper() for x in ["XH", "RIPPER", "W "])
                current = bit_me_map.get(bn)
                if current is None:
                    bit_me_map[bn] = (f, version, is_variant)
                else:
                    _, cur_ver, cur_variant = current
                    if is_variant and not cur_variant:
                        continue
                    if not is_variant and cur_variant:
                        bit_me_map[bn] = (f, version, is_variant)
                    elif version > cur_ver:
                        bit_me_map[bn] = (f, version, is_variant)

    matched = {bn: info[0] for bn, info in bit_me_map.items()}
    print(f"Matched {len(matched)} bits to Min Engagement files")

    # =========================================================================
    # Extract cutlet-based metrics for all matched bits
    # =========================================================================
    from cutlet_engine import read_cutter_data_from_me as read_me_for_cutlets

    all_metrics = {}
    for i, bit in enumerate(bits):
        bn = str(int(bit["bit_num"])) if isinstance(bit["bit_num"], (int, float)) else str(bit["bit_num"])
        if bn not in matched:
            continue

        filepath = matched[bn]
        print(f"  [{i+1}/{len(bits)}] Processing bit {bn}: {os.path.basename(filepath)}")

        try:
            # Extract cutter data for cutlet computation
            me_cutters, ipr_val, gage_rad = read_me_for_cutlets(filepath)

            # Compute cutlets via Shapely polygon operations
            cutlet_results = compute_cutlets(me_cutters, ipr_val, gage_rad)
            cutlet_results = [r for r in cutlet_results if r['area'] >= 0.001]

            # Also extract ME-file cutter data for orientation metrics (backrake, siderake)
            cutters_me, bit_diameter, engagement_thresholds = extract_cutters_from_me_file(filepath)
            if bit_diameter is None and bit.get("size"):
                bit_diameter = float(bit["size"])

            # Compute cutlet-derived metrics
            metrics = compute_cutlet_metrics(cutlet_results, cutters_me, gage_rad)
            if metrics:
                all_metrics[bn] = metrics
        except Exception as e:
            print(f"    ERROR: {e}")

    print(f"\nSuccessfully computed metrics for {len(all_metrics)} bits")

    if not all_metrics:
        print("No metrics computed. Exiting.")
        return

    bit_numbers_ordered = [bn for bn in [
        str(int(b["bit_num"])) if isinstance(b["bit_num"], (int, float)) else str(b["bit_num"])
        for b in bits
    ] if bn in all_metrics]

    # =========================================================================
    # GLOBAL SCORING — all bits scored on the same scale
    # =========================================================================
    # Previous approach (per-size-bucket with rescaling) amplified tiny metric
    # differences within small buckets, producing unintuitive 0-9 spreads.
    #
    # New approach: global min-max normalization across ALL bits.
    # All metrics used are already size-independent (ratios, CV, entropy,
    # density normalized by bit face area, etc.), so cross-size comparison
    # is valid.

    def score_global_0_9(components, comp_names, weights):
        """Score all bits globally from 0 to 9.

        1. Normalize each component across ALL bits to 0-1
        2. Weighted sum → raw score
        3. Scale raw scores to 0-9 globally
        Returns dict {bn: score}
        """
        bns = list(components.keys())
        if not bns:
            return {}

        # Normalize each component globally
        normed = {bn: dict(components[bn]) for bn in bns}
        for comp in comp_names:
            vals = [normed[bn][comp] for bn in bns]
            vmin, vmax = min(vals), max(vals)
            if vmax > vmin:
                for bn in bns:
                    normed[bn][comp] = (normed[bn][comp] - vmin) / (vmax - vmin)
            else:
                for bn in bns:
                    normed[bn][comp] = 0.5

        # Weighted sum
        raw = {}
        for bn in bns:
            raw[bn] = sum(normed[bn][comp] * weights[comp] for comp in comp_names)

        # Scale to 0-9
        rmin, rmax = min(raw.values()), max(raw.values())
        scores = {}
        for bn in bns:
            if rmax > rmin:
                scores[bn] = round(((raw[bn] - rmin) / (rmax - rmin)) * 9.0, 1)
            else:
                scores[bn] = 4.5
            scores[bn] = max(0.0, min(9.0, scores[bn]))
        return scores

    # =========================================================================
    # DURABILITY SCORING — cutlet-derived, global
    # =========================================================================
    # Physical basis: durability = how long the bit drills before pull.
    # 0 = most aggressive (wears fast), 9 = most durable (lasts long)
    #
    # Key physics:
    #   1. Cutter density (cutlets per sq inch of bit face): more cutlets
    #      sharing the same total rock volume = smaller individual cutlets
    #      = less force per cutter = more durable. Normalized by bit face
    #      area so it's comparable across bit sizes.
    #   2. Load balance (Gini): even per-cutter load = no single cutter
    #      overloaded = longer life.
    #   3. Chamfer fraction: 0.020" x 45° chamfer on all PDCs. Smaller
    #      cutlets have proportionally more chamfer = tougher cutting.
    #   4. Peak moderation: 1/max_mean_ratio — if the largest cutlet is
    #      much bigger than average, that cutter wears fastest.
    #   5. Blade balance (CV): even per-blade load distribution.
    #   6. Backrake: higher = more conservative cutting angle = less damage.
    #   7. Backup engagement: row 2 cutters forming cutlets = sharing load.
    #   8. Primary-secondary blade balance: when primary and secondary blade
    #      row 1 cutlets are similar size at overlapping radii, load is
    #      better distributed. Large differential = primaries doing most
    #      of the work = more aggressive.

    dur_comp_names = [
        "cutter_density",     # More cutlets per unit bit face = more durable
        "load_balance",       # 1 - Gini (even per-cutter load)
        "chamfer_toughness",  # Higher chamfer fraction = tougher cutters
        "peak_moderation",    # 1/max_mean_ratio (no single cutter overloaded)
        "blade_balance",      # 1 / (1 + blade_cv) (even per-blade load)
        "backrake",           # Conservative cutting angle
        "backup_engagement",  # Row 2 cutters forming cutlets = sharing load
        "pri_sec_balance",    # Primary-secondary blade cutlet size balance
    ]
    weights_dur = {
        "cutter_density": 0.22,      # Fundamental: more cutters = less force each
        "load_balance": 0.18,        # Per-cutter Gini — no single cutter overloaded
        "chamfer_toughness": 0.15,   # Chamfer physics: compression → toughness
        "peak_moderation": 0.12,     # Worst-case cutter overload
        "blade_balance": 0.10,       # No blade takes disproportionate damage
        "backrake": 0.10,            # Conservative cutting angle
        "backup_engagement": 0.07,   # Backups actively sharing load
        "pri_sec_balance": 0.06,     # Balanced primary-secondary blade work
    }

    durability_components = {}
    for bn in bit_numbers_ordered:
        m = all_metrics[bn]
        durability_components[bn] = {
            "cutter_density": m["cutter_density"],
            "load_balance": 1.0 - m["gini"],
            "chamfer_toughness": m["avg_chamfer_fraction"],
            "peak_moderation": 1.0 / max(m["max_mean_ratio"], 1.0),
            "blade_balance": 1.0 / (1.0 + m["blade_cv"]),
            "backrake": min(m["avg_backrake"] / 30.0, 1.0),
            "backup_engagement": m["backup_cutlet_fraction"],
            "pri_sec_balance": 1.0 / max(m["pri_sec_area_ratio"], 1.0),
        }

    durability_scores = score_global_0_9(durability_components, dur_comp_names, weights_dur)

    # =========================================================================
    # STEERABILITY SCORING — cutlet-derived, global
    # =========================================================================
    # Physical basis: steerability = how well the driller can control direction.
    # 0 = least steerable, 9 = most steerable
    #
    # Key physics from cutlet plots:
    #   1. Gauge area fraction: LESS gauge cutting = LESS wall grip = MORE
    #      steerable. This is the SHAPE of the force distribution at gauge.
    #   2. Gauge chamfer effect: gauge cutlets with high chamfer fraction are
    #      less aggressive laterally → easier to steer.
    #   3. Gauge backrake: conservative gauge (high BR) slides on wall = more
    #      steerable. Aggressive gauge (low BR) digs in = less steerable.
    #   4. Cone aggressiveness: lower cone backrake = builds angle better.
    #   5. Force concentration: concentrated force at nose/shoulder vs spread
    #      evenly. Concentrated = easier to pivot = more steerable.
    #   6. Blade factor: fewer blades = less gyroscopic stability.
    #   7. Siderake: directional force component aiding steerability.
    #   8. Profile depth (Y range): deeper profile = more side-cutting area.

    steer_comp_names = [
        "gauge_openness",         # 1 - gauge_area_fraction (shape-based)
        "gauge_chamfer_effect",   # Chamfer fraction at gauge (less aggressive wall cut)
        "gauge_conservatism",     # Gauge backrake (higher = slides more)
        "cone_aggressiveness",    # Lower cone BR = builds angle
        "force_concentration",    # Max bin fraction (concentrated = easier to pivot)
        "blade_factor",           # Fewer blades = more steerable
        "siderake",               # Directional force component
        "profile_depth",          # Deeper profile = more side-cutting
    ]
    weights_steer = {
        "gauge_openness": 0.25,        # Gauge contact is THE strongest driver
        "gauge_chamfer_effect": 0.12,  # Chamfer at gauge reduces aggressiveness
        "gauge_conservatism": 0.15,    # How aggressively gauge cutters dig in
        "cone_aggressiveness": 0.12,   # Aggressive cone builds angle
        "force_concentration": 0.10,   # Concentrated force = easier to pivot
        "blade_factor": 0.08,          # Fewer blades = less stable
        "siderake": 0.08,              # Directional force component
        "profile_depth": 0.10,         # Side-cutting capability
    }

    steerability_components = {}
    for bn in bit_numbers_ordered:
        m = all_metrics[bn]
        gauge_chamfer = m["weighted_chamfer_fraction"] * (1.0 + m["gauge_area_fraction"])

        steerability_components[bn] = {
            "gauge_openness": 1.0 - m["gauge_area_fraction"],
            "gauge_chamfer_effect": min(gauge_chamfer, 1.0),
            "gauge_conservatism": min(m["avg_gauge_backrake"] / 45.0, 1.0),
            "cone_aggressiveness": 1.0 - min(m["avg_cone_backrake"] / 30.0, 1.0),
            "force_concentration": m["max_bin_frac"],
            "blade_factor": 1.0 / max(m["num_blades"], 1),
            "siderake": m["avg_siderake"],
            "profile_depth": m["y_range"],
        }

    steerability_scores = score_global_0_9(steerability_components, steer_comp_names, weights_steer)

    # --- PRINT RESULTS ---
    print("\n" + "=" * 170)
    print(f"{'Bit':<7} {'Size':<6} {'Layout (ref)':<16} {'Cutlets':<8} {'Density':<8} "
          f"{'Gini':<6} {'ChamFr':<7} {'MaxMn':<6} {'BldCV':<6} {'BkpFr':<6} "
          f"{'PriSec':<7} {'BR':<6} {'GaugeFr':<8} "
          f"{'Dur':<6} {'Steer':<6}")
    print("=" * 170)
    for bit in bits:
        bn = str(int(bit["bit_num"])) if isinstance(bit["bit_num"], (int, float)) else str(bit["bit_num"])
        dur = durability_scores.get(bn, "-")
        steer = steerability_scores.get(bn, "-")
        layout = str(bit.get("layout_type", ""))[:14]
        m = all_metrics.get(bn)
        if m:
            print(f"  {bn:<7} {m['bit_diameter']:<6.2f} {layout:<16} {m['n_cutlets']:<8} "
                  f"{m['cutter_density']:<8.3f} "
                  f"{m['gini']:<6.3f} {m['avg_chamfer_fraction']:<7.3f} "
                  f"{m['max_mean_ratio']:<6.2f} {m['blade_cv']:<6.3f} "
                  f"{m['backup_cutlet_fraction']:<6.3f} "
                  f"{m['pri_sec_area_ratio']:<7.3f} {m['avg_backrake']:<6.1f} "
                  f"{m['gauge_area_fraction']:<8.3f} {dur:<6} {steer:<6}")
        else:
            print(f"  {bn:<7} {'?':<6} {layout:<16} {'':>8} {'':>8} {'':>6} {'':>7} {'':>6} {'':>6} {'':>6} {'':>7} {'':>6} {'':>8} {'-':<6} {'-':<6}")

    # --- WRITE TO WORKBOOK ---
    print(f"\nWriting scores to {WORKBOOK_PATH}...")
    wb_write = openpyxl.load_workbook(WORKBOOK_PATH)
    ws_write = wb_write["Sheet1"]

    updates = 0
    for row in ws_write.iter_rows(min_row=5, max_row=ws_write.max_row, values_only=False):
        bit_val = None
        ao_cell = ap_cell = None
        for c in row:
            if c.column_letter == "B": bit_val = c.value
            elif c.column_letter == "AO": ao_cell = c
            elif c.column_letter == "AP": ap_cell = c

        if bit_val is None:
            continue

        bn = str(int(bit_val)) if isinstance(bit_val, (int, float)) else str(bit_val)

        if bn in durability_scores and ao_cell is not None:
            ao_cell.value = durability_scores[bn]
            updates += 1
        elif ao_cell is not None:
            ao_cell.value = None

        if bn in steerability_scores and ap_cell is not None:
            ap_cell.value = steerability_scores[bn]
        elif ap_cell is not None:
            ap_cell.value = None

    wb_write.save(WORKBOOK_PATH)
    print(f"Updated {updates} rows in columns AO & AP")

    # Save detailed metrics to JSON
    metrics_output = {}
    for bn in bit_numbers_ordered:
        metrics_output[bn] = {
            "durability_score": durability_scores.get(bn),
            "steerability_score": steerability_scores.get(bn),
            "cutlet_metrics": {k: round(v, 4) if isinstance(v, float) else v
                               for k, v in all_metrics[bn].items()},
        }

    with open("bit_ratings_analysis.json", "w") as f:
        json.dump(metrics_output, f, indent=2)
    print("Saved detailed metrics to bit_ratings_analysis.json")


if __name__ == "__main__":
    main()
