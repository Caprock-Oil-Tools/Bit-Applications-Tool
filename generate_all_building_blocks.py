"""
Batch generate cutlet plots AND cutlet data files for all assemblies.
Single pass: reads source file once, computes polygons once, outputs both.
Skips outputs that already exist.
"""

import os
import sys
import traceback

PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, PROJECT_ROOT)

from cutlet_engine import read_cutter_data_from_me
from skills.bb2_cuttlet_plots import build_cutlet_polygons, plot_cutlets, get_blade_from_name, get_row_from_name

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

MASTER_DIR = os.path.join(PROJECT_ROOT, "Master")


def find_source_file(assy_dir):
    for f in os.listdir(assy_dir):
        if "Min Engagement" in f and f.endswith(".xlsm") and not f.startswith("~$"):
            return os.path.join(assy_dir, f)
    return None


def write_cutlet_data_excel(cutlet_polys, assy_name, ipr, gage_radius, output_path):
    """Write cutlet data (centroid X, Y, area) plus IPR to Excel."""
    # Extract data from polygons
    data = []
    for name, poly in cutlet_polys:
        if poly.is_empty or poly.area < 1e-10:
            continue
        centroid = poly.centroid
        data.append({
            'name': name,
            'blade': get_blade_from_name(name),
            'row': get_row_from_name(name),
            'centroid_x': round(-centroid.x, 4),
            'centroid_y': round(centroid.y, 4),
            'area': round(poly.area, 4),
        })
    data.sort(key=lambda r: r['centroid_x'])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cutlet Data"

    # Styles
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    num_align = Alignment(horizontal="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Title and parameters
    ws.merge_cells("A1:F1")
    ws["A1"] = f"{assy_name} \u2014 Cutlet Data"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = "IPR (in/rev):"
    ws["A2"].font = Font(bold=True, size=10)
    ws["B2"] = ipr
    ws["B2"].number_format = '0.000'
    ws["C2"] = "Gage Radius (in):"
    ws["C2"].font = Font(bold=True, size=10)
    ws["D2"] = gage_radius
    ws["D2"].number_format = '0.0000'

    # Headers
    headers = ["Name", "Blade", "Row", "Centroid X (in)", "Centroid Y (in)", "Area (in\u00b2)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    for i, d in enumerate(data, 5):
        values = [d['name'], d['blade'], d['row'], d['centroid_x'], d['centroid_y'], d['area']]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.alignment = num_align
            cell.border = thin_border
            if col >= 4:
                cell.number_format = '0.0000'

    # Summary
    summary_row = len(data) + 6
    ws.cell(row=summary_row, column=5, value="Count:").font = Font(bold=True)
    ws.cell(row=summary_row, column=5).alignment = Alignment(horizontal="right")
    ws.cell(row=summary_row, column=6, value=len(data)).font = Font(bold=True)

    total_row = summary_row + 1
    ws.cell(row=total_row, column=5, value="Total Area:").font = Font(bold=True)
    ws.cell(row=total_row, column=5).alignment = Alignment(horizontal="right")
    total_area = sum(d['area'] for d in data)
    cell = ws.cell(row=total_row, column=6, value=round(total_area, 4))
    cell.font = Font(bold=True)
    cell.number_format = '0.0000'

    # Column widths
    for col, w in enumerate([10, 8, 6, 16, 16, 14], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    wb.save(output_path)


def main():
    assemblies = sorted([
        d for d in os.listdir(MASTER_DIR)
        if os.path.isdir(os.path.join(MASTER_DIR, d))
    ])

    total = len(assemblies)
    plots_generated = 0
    data_generated = 0
    skipped = 0
    failed = []

    for i, assy_name in enumerate(assemblies, 1):
        assy_dir = os.path.join(MASTER_DIR, assy_name)

        source = find_source_file(assy_dir)
        if not source:
            continue

        # Read source to get IPR for filename construction
        try:
            cutters, ipr, gage_radius = read_cutter_data_from_me(source)
        except Exception as e:
            print(f"[{i}/{total}] {assy_name} — FAILED to read source: {e}")
            failed.append((assy_name, str(e)))
            continue

        file_stem = f"{assy_name} @ {ipr:.3f} IPR"
        plot_file = os.path.join(assy_dir, f"{file_stem}_cutlet_plot.png")
        data_file = os.path.join(assy_dir, f"{file_stem}_cutlet_data.xlsx")

        need_plot = not os.path.exists(plot_file)
        need_data = not os.path.exists(data_file)

        if not need_plot and not need_data:
            skipped += 1
            continue

        print(f"[{i}/{total}] {assy_name}", end="")
        if need_plot:
            print(" [plot]", end="")
        if need_data:
            print(" [data]", end="")
        print()

        try:
            cutlet_polys = build_cutlet_polygons(cutters, ipr, gage_radius)

            if need_plot:
                plot_cutlets(cutlet_polys, gage_radius, assy_name, ipr, plot_file)
                plots_generated += 1

            if need_data:
                write_cutlet_data_excel(cutlet_polys, assy_name, ipr, gage_radius, data_file)
                data_generated += 1
                print(f"  Data: {len(cutlet_polys)} cutlets")

        except Exception as e:
            print(f"  FAILED: {e}")
            traceback.print_exc()
            failed.append((assy_name, str(e)))

    print(f"\n{'='*60}")
    print(f"Plots generated: {plots_generated}")
    print(f"Data files generated: {data_generated}")
    print(f"Skipped (both exist): {skipped}")
    print(f"Failed: {len(failed)}")
    if failed:
        print("\nFailed assemblies:")
        for name, err in failed:
            print(f"  {name}: {err}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
